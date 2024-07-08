#MultuserbillingVAT
from django.shortcuts import render, redirect
from .models import *
from django.contrib import messages
from django.contrib.auth.models import auth
from django.utils.crypto import get_random_string
import random
from django.conf import settings
from django.core.mail import send_mail
from django.utils import timezone
from django.http.response import JsonResponse
from datetime import date
from django.db.models import F
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.views.generic import View
from io import BytesIO
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Max
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.views import View
import base64
from django.http import JsonResponse
from datetime import datetime,date, timedelta
from xhtml2pdf import pisa
from django.template.loader import get_template
from io import BytesIO
from django.template.loader import render_to_string
from reportlab.pdfgen import canvas
from django.core.files.base import ContentFile
from django.db import IntegrityError
from django.core.mail import EmailMultiAlternatives
from django.utils.html import strip_tags
from django.core.mail import send_mail, EmailMultiAlternatives
from django.http.response import JsonResponse, HttpResponse
from django.db.models import Sum,DecimalField
from django.shortcuts import redirect
from collections import defaultdict
from django.db.models import Sum, DecimalField
from .models import PurchaseBill, DebitNote, Party
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font,Protection
from django.db.models.functions import TruncMonth,TruncYear
import json
from django.db.models.functions import TruncMonth,TruncYear
import json
from calendar import month_name
from django.db.models import F, Sum, Value as V
import datetime

def home(request):
  return render(request, 'home.html')

def login(request):
  return render(request, 'login.html')

def forgot_password(request):
  return render(request, 'forgot_password.html')

def cmp_register(request):
  return render(request, 'cmp_register.html')

def cmp_details(request,id):
  context = {'id':id}
  return render(request, 'cmp_details.html', context)

def emp_register(request):
  return render(request, 'emp_register.html')

def register_company(request):
  if request.method == 'POST':
    fname = request.POST['fname']
    lname = request.POST['lname']
    email = request.POST['email']
    uname = request.POST['uname']
    phno = request.POST['phno']
    passw = request.POST['pass']
    cpass = request.POST['cpass']
    rfile = request.FILES.get('rfile')

    if passw == cpass:
      if CustomUser.objects.filter(username = uname).exists():
        messages.info(request, 'Sorry, Username already in Use !!')
        return redirect('cmp_register')
      
      elif Company.objects.filter(contact = phno).exists():
        messages.info(request, 'Sorry, Phone Number already in Use !!')
        return redirect('cmp_register')

      elif not CustomUser.objects.filter(email = email).exists():
        user_data = CustomUser.objects.create_user(first_name = fname, last_name = lname, username = uname, email = email, password = passw, is_company = 1)
        cmp = Company( contact = phno, user = user_data, profile_pic = rfile)
        cmp.save()
        return redirect('cmp_details',user_data.id)

      else:
        messages.info(request, 'Sorry, Email already in Use !!')
        return redirect('cmp_register')
      
    messages.info(request, 'Sorry, Passwords must match !!')
    return render(request,'cmp_register.html')
  
def register_company_details(request,id):
  if request.method == 'POST':
    cname = request.POST['cname']
    address = request.POST['address']
    city = request.POST['city']
    state = request.POST['state']
    country = request.POST['country']
    pincode = request.POST['pincode']
    pannumber = request.POST['pannumber']
    gsttype = request.POST['gsttype']
    gstno = request.POST['gstno']

    if Company.objects.filter(pan_number = pannumber).exclude(pan_number='').exists():
      messages.info(request, 'Sorry, Pan number is already in Use !!')
      return redirect('cmp_details',id)
    
    if Company.objects.filter(gst_no = gstno).exclude(gst_no='').exists():
      messages.info(request, 'Sorry, GST number is already in Use !!')
      return redirect('cmp_details',id)

    code=get_random_string(length=6)

    usr = CustomUser.objects.get(id = id)
    cust = Company.objects.get(user = usr)
    cust.company_name = cname
    cust.address = address
    cust.city = city
    cust.state = state
    cust.company_code = code
    cust.country = country
    cust.pincode = pincode
    cust.pan_number = pannumber
    cust.gst_type = gsttype
    cust.gst_no = gstno
    cust.save()
    return redirect('login')

def register_employee(request):
  if request.method == 'POST':
    fname = request.POST['fname']
    lname = request.POST['lname']
    email = request.POST['email']
    uname = request.POST['uname']
    phno = request.POST['phno']
    passw = request.POST['pass']
    cpass = request.POST['cpass']
    ccode = request.POST['ccode']
    rfile = request.FILES.get('rfile')

    if not Company.objects.filter(company_code = ccode).exists():
      messages.info(request, 'Sorry, Company Code is Invalid !!')
      return redirect('emp_register')
    
    cmp = Company.objects.get(company_code = ccode)
    emp_names = Employee.objects.filter(company = cmp).values_list('user',flat=True)
    for e in emp_names:
       usr = CustomUser.objects.get(id=e)
       if str(fname).lower() == (usr.first_name ).lower() and str(lname).lower() == (usr.last_name).lower():
        messages.info(request, 'Sorry, Employee With this name already exits, try adding an initial !!')
        return redirect('emp_register')
    
    if passw == cpass:
      if CustomUser.objects.filter(username = uname).exists():
        messages.info(request, 'Sorry, Username already exists !!')
        return redirect('emp_register')
      
      elif Employee.objects.filter(contact = phno).exists():
        messages.info(request, 'Sorry, Phone Number already in Use !!')
        return redirect('emp_register')

      elif not CustomUser.objects.filter(email = email).exists():
        user_data = CustomUser.objects.create_user(first_name = fname, last_name = lname, username = uname, email = email, password = passw)
        emp = Employee(user = user_data, company = cmp, profile_pic = rfile, contact=phno)
        emp.save()
        return redirect('login')

      else:
        messages.info(request, 'Sorry, Email already exists !!')
        return redirect('emp_register')
      
    messages.info(request, 'Sorry, Passwords must match !!')
    return render(request,'emp_register.html')
  
def change_password(request):
  if request.method == 'POST':
    email= request.POST.get('email')
    if not CustomUser.objects.filter(email=email).exists():
      messages.success(request,'Sorry, No user found with this email !!')
      return redirect('forgot_password')
    
    else:
      otp = random.randint(100000, 999999)
      usr = CustomUser.objects.get(email=email)
      usr.set_password(str(otp))
      usr.save()

      subject = 'Password Reset Mail'
      message = f'Hi {usr.first_name} {usr.last_name}, Your Otp for password reset is {otp}'
      email_from = settings.EMAIL_HOST_USER
      recipient_list = [email ]
      send_mail(subject, message, email_from, recipient_list)
      messages.info(request,'Password reset mail sent !!')
      return redirect('forgot_password')

def user_login(request):
  if request.method == 'POST':
    email = request.POST['email']
    cpass = request.POST['pass']

    try:
      usr = CustomUser.objects.get(email=email)
      log_user = auth.authenticate(username = usr.username, password = cpass)
      if log_user is not None:
        if usr.is_company == 1:
          auth.login(request, log_user)
          return redirect('dashboard')
        else:
          emp = Employee.objects.get(user=usr)
          if emp.is_approved == 0:
            messages.info(request,'Employee is not Approved !!')
            return redirect('login')
          else:
            auth.login(request, log_user)
            return redirect('dashboard')
      messages.info(request,'Invalid Login Details !!')
      return redirect('login')
    
    except:
        messages.info(request,'Employee do not exist !!')
        return redirect('login')
    

def dashboard(request):
  cmp =  Company.objects.get(user = request.user)
  current_time = datetime.datetime.now()
  month = 12
  total_invoice = Invoice.objects.filter(company_id=cmp).aggregate(total_invoice=Sum('subtotal'))['total_invoice'] or 0
  
  total_PurchaseBill = PurchaseBill.objects.filter(company_id=cmp).aggregate(total_PurchaseBill=Sum('grandtotal'))['total_PurchaseBill'] or 0

  total_PurchaseBill_jan = PurchaseBill.objects.filter(company_id=cmp,billdate__month=month).aggregate(total_PurchaseBill_jan=Sum('grandtotal'))['total_PurchaseBill_jan'] or 0

  
  context = {
     'usr':request.user,
     'total_invoice':total_invoice, 
     'total_PurchaseBill':total_PurchaseBill,
     'total_PurchaseBill_jan':total_PurchaseBill_jan,
     }
  
  return render(request, 'dashboard.html', context)

def logout(request):
  auth.logout(request)
  return redirect('/')

def cmp_profile(request):
  cmp = Company.objects.get(user = request.user)
  context = {'usr':request.user, 'cmp':cmp}
  return render(request,'cmp_profile.html',context)

def load_edit_cmp_profile(request):
  cmp = Company.objects.get(user = request.user)
  context = {'usr':request.user, 'cmp':cmp}
  return render(request,'cmp_profile_edit.html',context)

def edit_cmp_profile(request):
  cmp =  Company.objects.get(user = request.user)
  if request.method == 'POST':
    email = request.POST['email']
    current_email = cmp.user.email
    if email != current_email:
      if CustomUser.objects.filter(email=email).exists():
        messages.info(request,'Sorry, Email Already in Use !!')
        return redirect('load_edit_cmp_profile')
      
    phno_list = list(filter(None,Company.objects.exclude(user = request.user).values_list('contact', flat=True)))
    gst_list = list(filter(None,Company.objects.exclude(user = request.user).values_list('pan_number', flat=True)))
    gno_list = list(filter(None,Company.objects.exclude(user = request.user).values_list('gst_no', flat=True)))

    if request.POST['phno'] in phno_list:
      messages.info(request,'Sorry, Phone number already in Use !!')
      return redirect('load_edit_cmp_profile')

    if request.POST['pan'] in gst_list:
      messages.info(request,'Sorry, PAN number already in Use !!')
      return redirect('load_edit_cmp_profile')

    if request.POST['gstnoval'] in gno_list:
      messages.info(request,'Sorry, GST number already in Use !!')
      return redirect('load_edit_cmp_profile')

    cmp.company_name = request.POST['cname']
    cmp.user.email = request.POST['email']
    cmp.user.first_name = request.POST['fname']
    cmp.user.last_name = request.POST['lname']
    cmp.contact = request.POST['phno']
    cmp.address = request.POST['address']
    cmp.city = request.POST['city']
    cmp.state = request.POST['state']
    cmp.country = request.POST['country']
    cmp.pincode = request.POST['pincode']
    cmp.pan_number = request.POST['pan']
    cmp.gst_type = request.POST['gsttype']
    cmp.gst_no = request.POST['gstnoval']
    old=cmp.profile_pic
    new=request.FILES.get('image')
    if old!=None and new==None:
      cmp.profile_pic=old
    else:
      cmp.profile_pic=new
    
    cmp.save() 
    cmp.user.save() 
    return redirect('cmp_profile') 
  
def emp_profile(request):
  emp = Employee.objects.get(user=request.user)
  context = {'usr':request.user, 'emp':emp}
  return render(request,'emp_profile.html',context)

def load_edit_emp_profile(request):
  emp = Employee.objects.get(user=request.user)
  context = {'usr':request.user, 'emp':emp}
  return render(request,'emp_profile_edit.html',context)

def edit_emp_profile(request):
  emp =  Employee.objects.get(user = request.user)
  if request.method == 'POST':
    email = request.POST['email']
    current_email = emp.user.email
    if email != current_email:
      if CustomUser.objects.filter(email=email).exists():
        messages.info(request,'Email Already in Use')
        return redirect('load_edit_emp_profile')
          
    phno_list = list(Employee.objects.exclude(user = request.user).values_list('contact', flat=True))

    if request.POST['phno'] in phno_list:
      messages.info(request,'Sorry, Phone number already in Use !!')
      return redirect('load_edit_emp_profile')

    emp.user.email = request.POST['email']
    emp.user.first_name = request.POST['fname']
    emp.user.last_name = request.POST['lname']
    emp.contact = request.POST['phno']
    old=emp.profile_pic
    new=request.FILES.get('image')
    if old!=None and new==None:
      emp.profile_pic=old
    else:
      emp.profile_pic=new
    
    emp.save() 
    emp.user.save() 
    return redirect('emp_profile') 

def load_staff_request(request):
  cmp = Company.objects.get(user = request.user)
  emp = Employee.objects.filter(company = cmp, is_approved = 0)
  context = {'usr':request.user, 'emp':emp, 'cmp':cmp}
  return render(request,'staff_request.html',context)

def load_staff_list(request):
  cmp = Company.objects.get(user = request.user)
  emp = Employee.objects.filter(company = cmp, is_approved = 1)
  context = {'usr':request.user, 'emp':emp, 'cmp':cmp}
  return render(request,'staff_list.html',context)

def accept_staff(request,id):
  emp = Employee.objects.get(id=id)
  emp.is_approved = 1
  emp.save()
  messages.info(request,'Employee Approved !!')
  return redirect('load_staff_request')

def reject_staff(request,id):
  emp = Employee.objects.get(id=id)
  emp.user.delete()
  emp.delete()
  messages.info(request,'Employee Deleted !!')
  return redirect('load_staff_request')

def item_list_first(request):
  if request.user.is_company:
    itm_list = Item.objects.filter(company = request.user.company)
  else:
    itm_list = Item.objects.filter(company = request.user.employee.company)
  
  if itm_list:
    itm = itm_list[0]
    trans = ItemTransactions.objects.filter(item = itm.id)
    print(itm)
    context = {'itm_list':itm_list, 'usr':request.user, 'itm':itm, 'trans':trans}
  else:
    context = {'itm_list':itm_list, 'usr':request.user}
  return render(request,'item_list.html',context)

def item_list(request,id):
  if request.user.is_company:
    itm_list = Item.objects.filter(company = request.user.company)
  else:
    itm_list = Item.objects.filter(company = request.user.employee.company)
  
  itm = Item.objects.get(id=id)
  trans = ItemTransactions.objects.filter(item=itm.id)
  bal_quantity = ItemTransactions.objects.filter(item=itm.id).last().trans_current_qty
  
  context = {'itm_list':itm_list, 'usr':request.user, 'itm':itm, 'trans':trans,'bal_quantity':bal_quantity}
  return render(request,'item_list.html',context)  

def load_item_create(request):
  tod = timezone.now().date().strftime("%Y-%m-%d")
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  unit = Unit.objects.filter(company=cmp)
  return render(request,'item_create.html',{'tod':tod, 'usr':request.user, 'unit':unit})

def item_create(request):
  if request.method=='POST':
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company
  
    itm_type = request.POST.get('itm_type')
    itm_name = request.POST.get('name')
    itm_hsn = request.POST.get('hsn')

    if Item.objects.filter(company=cmp,itm_name=itm_name).exists():
      messages.info(request,'Item already registered')
      return redirect('load_item_create')

    if Item.objects.filter(company=cmp,itm_hsn=itm_hsn).exists():
      messages.info(request,'Item with this HSN already registered')
      return redirect('load_item_create')

    itm_unit = request.POST.get('unit')
    itm_vat = request.POST.get('vat')
    taxable_result = request.POST.get('taxable_result')
    itm_sale_price = request.POST.get('sale_price')
    itm_purchase_price = request.POST.get('purchase_price')
    stock_in_hand = request.POST.get('stock_in_hand')
    if stock_in_hand == '' or None :
      stock_in_hand = 0
    itm_at_price = request.POST.get('at_price')
    if itm_at_price == '' or None:
      itm_at_price = 0
    itm_date = request.POST.get('itm_date')
    
    item = Item(user = request.user,
                itm_type = itm_type,
                itm_name = itm_name,
                itm_hsn = itm_hsn,
                itm_unit = itm_unit,
                itm_vat = itm_vat,
                itm_taxable = taxable_result,
                itm_sale_price = itm_sale_price,
                itm_purchase_price = itm_purchase_price,
                itm_stock_in_hand = stock_in_hand,
                itm_at_price = itm_at_price,
                itm_date = itm_date)
    item.save()

    trans = ItemTransactions(user = request.user, item = item, trans_type = 'Stock Open', trans_date = itm_date, trans_qty = stock_in_hand, 
                             trans_current_qty = stock_in_hand, trans_adjusted_qty = stock_in_hand, trans_price = itm_at_price)

    if request.user.is_company:
      item.company = request.user.company
      trans.company = request.user.company

    else:
      item.company = request.user.employee.company
      trans.company = request.user.employee.company
 
    item.save()
    trans.save()

    trhis = ItemTransactionsHistory(user = request.user, transaction = trans, hist_trans_qty = 0, hist_trans_current_qty = stock_in_hand, action = 'Created',
                                    hist_trans_adjusted_qty = stock_in_hand, hist_trans_type = trans.trans_type)
    trhis.save()

    if request.POST.get('save_and_next'):
      return redirect('load_item_create')
    elif request.POST.get('save'):
      return redirect('item_list_first')
    
def adjust_stock(request,id):
  if request.method=='POST':
    itm = Item.objects.get(id=id)
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company

    trans_type = request.POST.get('trans_type')
    if trans_type == 'on':
      trans_type = 'Stock Reduction'
      trans_qty = request.POST.get('reduced_qty')
    else:
      trans_type = 'Stock Addition'
      trans_qty = request.POST.get('added_qty')
    trans_date = request.POST.get('trans_date')

    adjusted_qty= request.POST.get('adjusted_qty')
    current_qty = request.POST.get('item_qty')
    itm.itm_stock_in_hand = adjusted_qty
    itm.save()
    trans = ItemTransactions(user=request.user,
                          company=cmp,
                          item=itm,
                          trans_type=trans_type,
                          trans_date=trans_date,
                          trans_qty=trans_qty,
                          trans_current_qty=current_qty,
                          trans_adjusted_qty=adjusted_qty)
    trans.save()

    trhis = ItemTransactionsHistory(user = request.user, transaction = trans, hist_trans_qty = trans_qty, hist_trans_current_qty = current_qty, action = 'Created',
                                    hist_trans_adjusted_qty = adjusted_qty, hist_trans_type = trans.trans_type)
    trhis.save()

  return redirect('item_list',id)

@csrf_exempt
def create_unit(request):
    if request.method == 'POST':
        unit_name = request.POST.get('unit_name', '')
        company_id = request.user.company.id
        print(f"Company ID: {company_id}")

        company = Company.objects.get(id=request.user.company.id)

        # Create a new Unit instance
        unit = Unit.objects.create(
            company=company,
            unit_name=unit_name
        )
        
        # Prepare the JSON response data
        response_data = {
            'success': True,
            'message': 'Unit created successfully!',
            'unit_name': unit.unit_name,
            'unit_id': unit.id
        }
        print("Response Data:", response_data)

        return JsonResponse(response_data)

    # Handle other HTTP methods if needed
    return JsonResponse({'success': False, 'message': 'Invalid request method'})
  
def load_item_edit(request,id):
  itm = Item.objects.get(id=id)
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  unit = Unit.objects.filter(company=cmp)
  context = {'usr':request.user, 'itm':itm, 'unit':unit}
  return render(request,'item_edit.html',context)
  
def item_edit(request,id):
  if request.method == 'POST':
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company

    itm = Item.objects.get(id=id)
    itm.itm_type = request.POST.get('itm_type')
    itm.itm_name = request.POST.get('name')
    itm.itm_hsn = request.POST.get('hsn')

    itm_name = request.POST.get('name')
    itm_hsn = request.POST.get('hsn')
    
    if Item.objects.filter(company=cmp,itm_name=itm_name).exclude(itm_name = itm.itm_name).exists():
      messages.info(request,'Item already registered')
      return redirect('load_item_create')

    if Item.objects.filter(company=cmp,itm_hsn=itm_hsn).exclude(itm_hsn = itm.itm_hsn).exists():
      messages.info(request,'Item with this HSN already registered')
      return redirect('load_item_create')

    itm.itm_unit = request.POST.get('unit')
    itm.itm_vat = request.POST.get('vat')
    itm.itm_taxable = request.POST.get('taxable_result')
    itm.itm_sale_price = request.POST.get('sale_price')
    itm.itm_purchase_price = request.POST.get('purchase_price')

    stock_in_hand = request.POST.get('stock_in_hand')
    if stock_in_hand == '' or None :
      stock_in_hand = 0
    itm.itm_stock_in_hand = stock_in_hand

    itm_at_price = request.POST.get('at_price')
    if itm_at_price == '' or None:
      itm_at_price = 0
    itm.itm_at_price = itm_at_price

    itm.itm_date = request.POST.get('itm_date')
    itm.save()

    etrans = ItemTransactions.objects.filter(item=itm)
    new_stock_in_hand = stock_in_hand
    for e in etrans:
      e.trans_current_qty = new_stock_in_hand
      if e.trans_type == 'Stock Open':
        stock_val = int(new_stock_in_hand) + 0
      elif e.trans_type == 'Stock Addition':
        stock_val = int(e.trans_qty) + int(new_stock_in_hand)
      else:
        stock_val = int(new_stock_in_hand) - int(e.trans_qty)

      e.trans_adjusted_qty = stock_val
      e.save()
      
      trhis = ItemTransactionsHistory(user = request.user, transaction = e, hist_trans_current_qty = new_stock_in_hand, action = 'Updated',
                                        hist_trans_adjusted_qty = stock_val, hist_trans_type = e.trans_type)
      
      if e.trans_type == 'Stock Open':
        trhis.hist_trans_qty = 0
      else:
        trhis.hist_trans_qty = e.trans_qty
      trhis.save()
      new_stock_in_hand = stock_val

    itm.itm_stock_in_hand = new_stock_in_hand
    itm.save()

    if request.POST.get('save_and_next'):
      return redirect('load_item_create')
    elif request.POST.get('save'):
      return redirect('item_list',id)
    
  
def load_first_trans_edit(request,id):
  trans = ItemTransactions.objects.get(id=id)
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  unit = Unit.objects.filter(company=cmp)
  context = {'usr':request.user, 'trans':trans, 'unit':unit}
  return render(request,'item_first_trans_edit.html',context)
  
def first_trans_edit(request,id):
  if request.method == 'POST':
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company

    itm = Item.objects.get(id=id)
    itm.itm_type = request.POST.get('itm_type')
    itm.itm_name = request.POST.get('name')
    itm.itm_hsn = request.POST.get('hsn')

    itm_name = request.POST.get('name')
    itm_hsn = request.POST.get('hsn')
    
    if Item.objects.filter(company=cmp,itm_name=itm_name).exclude(itm_name = itm.itm_name).exists():
      messages.info(request,'Item already registered')
      return redirect('load_item_create')

    if Item.objects.filter(company=cmp,itm_hsn=itm_hsn).exclude(itm_hsn = itm.itm_hsn).exists():
      messages.info(request,'Item with this HSN already registered')
      return redirect('load_item_create')

    itm.itm_unit = request.POST.get('unit')
    itm.itm_vat = request.POST.get('vat')
    itm.itm_taxable = request.POST.get('taxable_result')
    itm.itm_sale_price = request.POST.get('sale_price')
    itm.itm_purchase_price = request.POST.get('purchase_price')

    stock_in_hand = request.POST.get('stock_in_hand')
    if stock_in_hand == '' or None :
      stock_in_hand = 0
    itm.itm_stock_in_hand = stock_in_hand

    itm_at_price = request.POST.get('at_price')
    if itm_at_price == '' or None:
      itm_at_price = 0
    itm.itm_at_price = itm_at_price

    itm.itm_date = request.POST.get('itm_date')
    itm.save()

    etrans = ItemTransactions.objects.filter(item=itm)
    new_stock_in_hand = stock_in_hand
    for e in etrans:
      e.trans_current_qty = new_stock_in_hand
      if e.trans_type == 'Stock Open':
        stock_val = int(new_stock_in_hand) + 0
      elif e.trans_type == 'Stock Addition':
        stock_val = int(e.trans_qty) + int(new_stock_in_hand)
      else:
        stock_val = int(new_stock_in_hand) - int(e.trans_qty)

      e.trans_adjusted_qty = stock_val
      e.save()
      
      trhis = ItemTransactionsHistory(user = request.user, transaction = e, hist_trans_current_qty = new_stock_in_hand, action = 'Updated',
                                        hist_trans_adjusted_qty = stock_val, hist_trans_type = e.trans_type)
      
      if e.trans_type == 'Stock Open':
        trhis.hist_trans_qty = 0
      else:
        trhis.hist_trans_qty = e.trans_qty
      trhis.save()
      new_stock_in_hand = stock_val

    itm.itm_stock_in_hand = new_stock_in_hand
    itm.save()

    if request.POST.get('save_and_next'):
      return redirect('load_item_create')
    elif request.POST.get('save'):
      return redirect('item_list',id)
    
def load_trans_details(request):
  id = request.POST.get('id')
  trans = ItemTransactions.objects.get(id=id)
  trans_id = trans.id
  name = trans.item.itm_name
  date = trans.trans_date.strftime("%Y-%m-%d")
  current_qty = trans.trans_current_qty
  qty = trans.trans_qty
  adj_qty = trans.trans_adjusted_qty
  trans_type = trans.trans_type
  return JsonResponse({'message': 'success',
                       'trans_id':trans_id,
                       'name':name, 
                       'date':date,
                       'current_qty':current_qty, 
                       'qty':qty, 
                       'adj_qty':adj_qty, 
                       'trans_type':trans_type})

def edit_transactions(request):
  if request.method=='POST':
    id = request.POST.get('id')
    trans_stock_change = request.POST.get('trans_stock_change')
    trans_itm_date = request.POST.get('trans_itm_date')
    trans_item_quantity = request.POST.get('trans_item_quantity')
    stock_value = request.POST.get('stock_value')
    trans_adj_quantity = request.POST.get('trans_adj_quantity')

    trans = ItemTransactions.objects.get(id=id)
    trans.trans_type = trans_stock_change
    trans.trans_date = trans_itm_date
    trans.trans_qty = stock_value 
    trans.trans_current_qty = trans_item_quantity
    trans.trans_adjusted_qty = trans_adj_quantity
    trans.save()

    trhis = ItemTransactionsHistory(user = request.user, transaction = trans, hist_trans_qty = stock_value, hist_trans_current_qty = trans_item_quantity, 
                                    action = 'Updated',hist_trans_adjusted_qty = trans_adj_quantity, hist_trans_type = trans.trans_type)
    trhis.save()

    etrans = ItemTransactions.objects.filter(item = trans.item,id__gt=id)
    new_stock_in_hand = trans_adj_quantity
    for e in etrans:
      e.trans_current_qty = new_stock_in_hand
      if e.trans_type == 'Stock Open':
        stock_val = int(new_stock_in_hand) + 0
      elif e.trans_type == 'Stock Addition':
        stock_val = int(e.trans_qty) + int(new_stock_in_hand)
      else:
        stock_val = int(new_stock_in_hand) - int(e.trans_qty)

      e.trans_adjusted_qty = stock_val
      e.save()
      
      trhis = ItemTransactionsHistory(user = request.user, transaction = e, hist_trans_current_qty = new_stock_in_hand, action = 'Updated',
                                        hist_trans_adjusted_qty = stock_val, hist_trans_type = e.trans_type)
      
      if e.trans_type == 'Stock Open':
        trhis.hist_trans_qty = 0
      else:
        trhis.hist_trans_qty = e.trans_qty
      trhis.save()
      new_stock_in_hand = stock_val

    trans.item.itm_stock_in_hand = new_stock_in_hand
    trans.item.save()

    return JsonResponse({'message': 'success', 'id':trans.item.id})

def delete_item(request,id):
  Item.objects.get(id=id).delete()
  messages.info(request,'Item Deleted Successfully !!')
  return redirect('item_list_first')

def delete_transaction(request,id):
  dtrans = ItemTransactions.objects.get(id=id)
  itm = dtrans.item
  dtrans.delete()

  trans = ItemTransactions.objects.filter(item=itm)
  cstock = trans[0].trans_adjusted_qty
  for t in range(1,len(trans)):
    trans[t].trans_current_qty = cstock
    if trans[t].trans_type == 'Stock Addition':
      trans[t].trans_adjusted_qty = trans[t].trans_qty + cstock
      fstock = trans[t].trans_qty + cstock
    else:
      trans[t].trans_adjusted_qty = cstock - trans[t].trans_qty
      fstock = cstock - trans[t].trans_qty
    trans[t].save()

    if trans[t].id > id:
      trhis = ItemTransactionsHistory(user = request.user, transaction = trans[t], hist_trans_qty = trans[t].trans_qty, hist_trans_current_qty = cstock, 
                                      action = 'Updated', hist_trans_adjusted_qty = fstock, hist_trans_type = trans[t].trans_type)
      trhis.save()
    cstock = fstock

  itm.itm_stock_in_hand = cstock
  itm.save()

  messages.info(request,'Item Transaction Deleted Successfully !!')
  return redirect('item_list_first')

def load_itm_trans_history(request, id):
    if request.user.is_company:
        itm= Item.objects.filter(company=request.user.company)
    else:
        itm = Item.objects.filter(company=request.user.employee.company)

    fitem = Item.objects.get(id=id)
    #fitems = ItemTransactions.objects.get(item=fitem)
    hst = ItemTransactionsHistory.objects.filter(transaction=id)

    context = {'itm': itm, 'hst': hst, 'usr': request.user, 'fitem': fitem}
    return render(request, 'item_transaction_history.html',context)
  
  
def allbill(request):
    if request.user.is_company:
          cmp = request.user.company
    else:
          cmp = request.user.employee.company
    usr = CustomUser.objects.get(username=request.user) 
   
    itm=PurchaseBill.objects.filter(company=cmp)
    pbill = PurchaseBill.objects.filter(company=cmp).values()
    pbills = PurchaseBill.objects.filter(company=cmp)

    for i in pbill:
      p_history= PurchaseBillTransactionHistory.objects.filter(purchasebill=i['id'],company=cmp).last()
      if p_history:
        i['action'] = p_history.action
        i['name']= f"{p_history.staff.first_name} {p_history.staff.last_name}"
        i['party'] = p_history.purchasebill.party.party_name
      else:
        # Handle the case when no history is found
        i['action'] = ""
        i['name'] = ""
        i['party_name'] = ""
    
      # i['action']=p_history.action
      # i['name']=p_history.staff.first_name+" "+p_history.staff.last_name
      # i['party']=p_history.purchasebill.party.party_name
    return render(request, 'all_billdetils.html',{'itm':itm,'pbill':pbill,'pbills':pbills,'usr':request.user})
    
def purchasebill(request):
    if request.user.is_company:
          cmp = request.user.company
    else:
          cmp = request.user.employee.company
    usr = CustomUser.objects.get(username=request.user) 
    party=Party.objects.filter(company=cmp)
    item=Item.objects.filter(company=cmp)
    unit=Unit.objects.filter(company=cmp)
    last_bill = PurchaseBill.objects.filter(company=cmp).order_by('-billno').first()
    if last_bill:
        bill_no = last_bill.billno + 1
    else:
        bill_no = 1
    # print (bill_no)
    context = {
        'bill_no': bill_no,
        'party':party,
        'item':item,
        'unit':unit,
        'usr':request.user
        # Add other context variables as needed
    }
    return render(request, 'createpurchasebill.html',context)
    
def itemdetails(request):
  itmid = request.GET['id']
  itm = Item.objects.get(id=itmid)
  hsn = itm.itm_hsn
  vat = itm.itm_vat
  price = itm.itm_purchase_price
  qty = 0
  print(qty)
  return JsonResponse({'hsn':hsn, 'vat':vat,  'price':price, 'qty':qty})
  
def item_dropdown(request):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company
  options = {}
  option_objects = Item.objects.filter(company=cmp)
  for option in option_objects:
      options[option.id] = [option.id, option.itm_name]
  return JsonResponse(options)

def cust_dropdown(request):
    if request.user.is_company:
          cmp = request.user.company
    else:
          cmp = request.user.employee.company
    usr = CustomUser.objects.get(username=request.user) 
    party=Party.objects.filter(company=cmp,user=usr)

    id_list = []
    party_list = []
    for p in party:
      id_list.append(p.id)
      party_list.append(p.party_name)

    return JsonResponse({'id_list':id_list, 'party_list':party_list })


def custdata(request):
  cid = request.POST['id']
  part = Party.objects.get(id=cid)
  phno = part.contact
  address = part.address
  pay = part.payment
  bal = part.openingbalance
  return JsonResponse({'phno':phno, 'address':address, 'pay':pay, 'bal':bal})
  
def createbill(request):
  if request.method == 'POST':
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company
    usr = CustomUser.objects.get(username=request.user)
    
    # reference_no=request.POST.get('refnum')
    # print("reference_no", reference_no)
    part = Party.objects.get(id=request.POST.get('customername'))
    print("part",part)
    billno=request.POST.get('bill_no')
    print("billno",billno)
    billdate=request.POST.get('billdate')
    print("billdate",billdate)
    subtotal=float(request.POST.get('subtotal'))
    print("subtotal",subtotal)
    adjust = request.POST.get("adj")
    print("adjust",adjust)
    taxamount = request.POST.get("taxamount")
    print("taxamount",taxamount)
    grandtotal=request.POST.get('grandtotal')
    print("grandtotal",grandtotal)
    
    
    pbill = PurchaseBill(
      staff=usr,
      company=cmp,
      party=part, 
      billno=billno, 
      billdate=billdate,
      subtotal=subtotal,
      taxamount=taxamount,
      adjust=adjust, 
      grandtotal=grandtotal)
    pbill.save()
    print("pbill: ",pbill)
    
    product =request.POST.getlist('product[]')
    print("product: ",product)
    qty =request.POST.getlist('qty[]')
    print("item qty: ",qty)
    price =request.POST.getlist('price[]')
    print("item price",price)
    tax =request.POST.getlist('vat[]')
    print("item tax",tax)
    discount = request.POST.getlist('discount[]')
    print("item discount",discount)
    hsn = request.POST.getlist('hsn[]')
    print("item hsn",hsn)
    total = request.POST.getlist('total[]')
    print("item total",total)
    
    if len(product) == len(qty) == len(price) == len(tax) == len(discount) == len(hsn) == len(total) and product and qty and price and tax and discount and hsn and total:
      mapped=zip(product,qty,price,tax,discount,hsn,total)
      mapped=list(mapped)
      for ele in mapped:
        print("Element:", ele)
        hsn = ele[5]
        qty = ele[1]
        tax = ele[3]
        price = ele[2]
        discount = ele[4]
        total = ele[6]
        print("HSN:", hsn)
        print("Quantity:", qty)
        print("Tax:", tax)
        print("Price:", price)
        print("Discount:", discount)
        print("Total:", total)

        item_name_parts = ele[0].split()
        print("item_name_parts: ",item_name_parts)
        item_id = item_name_parts[0]
        print("item_id: ",item_id)
        items = Item.objects.get(id=item_id,company=cmp)
        print("items are: ",items)
        it=Item.objects.get(company=cmp, id = item_id).itm_name
        print("item_name:", it)
        
        purchasebillitem=PurchaseBillItem(
                  user=usr,
                  purchasebill=pbill,
                  company=cmp,
                  product=items,
                  item=it,
                  hsn=ele[5],
                  qty=ele[1],
                  VAT=ele[3],
                  price=ele[2],
                  discount=ele[4],
                  total=ele[6])
        purchasebillitem.save()
        print("purchasebillitem:", purchasebillitem)
    
    itt = ItemTransactions.objects.filter(item=items.id).last().trans_current_qty
    # Add invoice details in items transactions
    transaction = ItemTransactions.objects.create(
        company=cmp,
        item=items,
        trans_type='Purchase',
        trans_date=pbill.billdate,
        trans_qty=qty,
        trans_current_qty = itt + int(qty),
        trans_price=price,
        #trans_invoice=invoice_curr.invoice_no
    )
    items.itm_stock_in_hand+= int(ele[1])
    items.save()
    print("Transaction saved:", transaction)
    
    # p_id  = Party.objects.get(id=request.POST.get('partyname'))
    print('transaction0')
    part.openingbalance=float(part.openingbalance)-float(pbill.grandtotal)
    part.save()
    transactionparty=Transactions_party.objects.create(
      company=cmp,
      user=usr,
      party=part,
      trans_type='Purchase',
      trans_number=pbill.billno,
      trans_date=pbill.billdate,
      total=pbill.grandtotal,
      balance=part.openingbalance,
      
    )
    print(transactionparty.party)   
    print("Transaction party1:", transactionparty)
    transactionparty.save()
      
    cr=PurchaseBillTransactionHistory(staff=usr,company=cmp,purchasebill=pbill,action='Created')
    print("cr", cr)
    cr.save()
    if 'Next' in request.POST:
        return redirect('purchasebill')
    if "Save" in request.POST:
        return redirect('allbill') 
  else:
      return redirect('purchasebill')
     
def billhistory(request):
  pid = request.POST['id']
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
  pbill = PurchaseBill.objects.get(billno=pid,company=cmp,staff=usr)
  hst = PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp).last()
  name = hst.staff.first_name + ' ' + hst.staff.last_name 
  action = hst.action
  return JsonResponse({'name':name,'action':action,'pid':pid,'usr':request.user})


def delete_purchasebill(request,id):
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
  pbill = PurchaseBill.objects.get(id=id)
  PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp).delete()
  pbill.delete()
  
  return redirect('allbill')



def details_purchasebill(request,id):
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  pitm = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)
  dis = 0
  for itm in pitm:
    dis += int(itm.discount)
  itm_len = len(pitm)

  context={'pbill':pbill,'pitm':pitm,'itm_len':itm_len,'dis':dis,'usr':request.user}
  return render(request,'vatbilldetils.html',context)
  
def history_purchasebill(request,id):
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company 
  usr = CustomUser.objects.get(username=request.user) 
  pbill = PurchaseBill.objects.get(id=id)
  hst= PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp)

  context = {'hst':hst,'pbill':pbill,'usr':request.user}
  return render(request,'purchasebillhistory.html',context)
  
def edit_purchasebill(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
  cust = Party.objects.filter(company=cmp)
  item = Item.objects.filter(company=cmp)
  item_units = Unit.objects.filter(company=cmp)

  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  billprd = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)
  bdate = pbill.billdate.strftime("%Y-%m-%d")
  context = { 'pbill':pbill, 'billprd':billprd,'tod':tod,
             'cust':cust, 'item':item, 'item_units':item_units, 'bdate':bdate,'usr':request.user}
  return render(request,'purchasebilledit.html',context)
  
def save_purchasebill(request,id):
  if request.method =='POST':
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company  

    usr = CustomUser.objects.get(username=request.user) 
    print('haiii')
    print (request.POST.get('customername'))
    part = Party.objects.get(id=request.POST.get('customername'))
    print(part)
    pbill = PurchaseBill.objects.get(id=id, company=cmp)
    print(pbill)

        # Access the related PurchaseBill instance through the ForeignKey


    pbill.party = part
    pbill.billdate = request.POST.get('billdate')
    pbill.subtotal =float(request.POST.get('subtotal'))
    pbill.grandtotal = request.POST.get('grandtotal')
    pbill.taxamount = request.POST.get("taxamount")
    pbill.adjust = request.POST.get("adj")
    pbill.company=cmp
    pbill.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    tax =  tuple(request.POST.getlist("vat[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    PurchaseBillItem.objects.filter(purchasebill=pbill).delete()
    if len(total)==len(discount)==len(qty)==len(tax):
      mapped=zip(product,qty,tax,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = Item.objects.get(id=ele[0])
        PurchaseBillItem.objects.create(product =itm,qty=ele[1], VAT=ele[2],discount=ele[3],total=ele[4],purchasebill=pbill,company=cmp)

    PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,action='Updated',company=cmp,staff=usr)
    return redirect('allbill')

  return redirect('allbill')

def save_unit(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    usr = CustomUser.objects.get(username=request.user) 

    if request.method == 'POST':
      unit_name = request.POST.get('name')
      if unit_name == 'NUMBERS' or unit_name == 'BOX' or unit_name == 'PACKETS' or unit_name == 'numbers' or unit_name == 'box'  or unit_name == 'packets':
        return JsonResponse({'success': False})
      if Unit.objects.filter(unit_name=unit_name,company=cmp).exists():
         return JsonResponse({'success': False})
      itm = Unit(
          company=cmp,
          unit_name=unit_name,
       
      )
      itm.save()

      return JsonResponse({'success': True})
    return JsonResponse({'success': False})
    
def save_item(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    usr = CustomUser.objects.get(username=request.user) 

    if request.method == 'POST':
      itm_type = request.POST.get('itm_type')
      name = request.POST.get('name')
      itm_hsn = request.POST.get('hsn')
      itm_unit = request.POST.get('unit')
      itm_taxable = request.POST.get('taxref')
      itm_vat = request.POST.get('vat')
      itm_sale_price = request.POST.get('sell_price')
      itm_purchase_price = request.POST.get('cost_price')
      itm_stock_in_hand = request.POST.get('stock ', 0)  # Default to 0 if not provided
      itm_at_price = request.POST.get('itmprice ', 0)  # Default to 0 if not provided
      itm_date = request.POST.get('itmdate')

      # Check if the HSN number already exists in the database
      if Item.objects.filter(itm_hsn=itm_hsn,company=cmp).exists():
          # Send a message indicating that the HSN number already exists
          return JsonResponse({'success': False})

      itm = Item(
          user=usr,
          company=cmp,
          itm_type=itm_type,
          itm_name=name,
          itm_hsn=itm_hsn,
          itm_unit=itm_unit,
          itm_taxable=itm_taxable,
          itm_vat=itm_vat,
          itm_sale_price=itm_sale_price,
          itm_purchase_price=itm_purchase_price,
          itm_stock_in_hand=itm_stock_in_hand,
          itm_at_price=itm_at_price,
          itm_date=itm_date
      )
      itm.save()

      return JsonResponse({'success': True})
    return JsonResponse({'success': False})
        
def save_party1(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    usr = CustomUser.objects.get(username=request.user)

    if request.method == 'POST':
        partyname = request.POST.get('partyname')
        trn_no = request.POST.get('trn_no')
        contact = request.POST.get('contact')
        placeof=request.POST.get('pol')
        trn_type = request.POST.get('trn_type')
        address = request.POST.get('address')
        email = request.POST.get('email')
        balance = request.POST.get('balance')
        paymentType = request.POST.get('paymentType')
        currentdate = request.POST.get('currentdate')
        additionalfield1 = request.POST.get('additionalfield1')
        additionalfield2 = request.POST.get('additionalfield2')
        additionalfield3 = request.POST.get('additionalfield3')
        # print(trn_no)
        # print(partyname)

        # Check if the contact number already exists in the database
        if Party.objects.filter(contact=contact,company=cmp).exists():
            # Send a message indicating that the contact number already exists

            return redirect('createbill')

        # Check if the transaction number already exists in the database
        if trn_type == "Unregistered/Consumers":
            Party.objects.create(    user=usr,
            company=cmp,
            party_name=partyname,
            trn_no=trn_no,
            contact=contact,
            trn_type=trn_type,
            address=address,
            email=email,
            state=placeof,
            openingbalance=balance,
            payment=paymentType,
            current_date=currentdate,
            additionalfield1=additionalfield1,
            additionalfield2=additionalfield2,
            additionalfield3=additionalfield3)
            # Optionally, you can send a success message here

            return redirect('createbill')
        else:
             if Party.objects.filter(trn_no=trn_no, company=cmp).exists():
                # Send a message indicating that the transaction number already exists

                return redirect('createbill')
        Party.objects.create(
            user=usr,
            company=cmp,
            party_name=partyname,
            trn_no=trn_no,
            contact=contact,
            trn_type=trn_type,
            address=address,
            email=email,
            state=placeof,
            openingbalance=balance,
            payment=paymentType,
            current_date=currentdate,
            additionalfield1=additionalfield1,
            additionalfield2=additionalfield2,
            additionalfield3=additionalfield3
        )
        return redirect('createbill')
        
def sharepdftomail(request,id):
 if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                if request.user.is_company:
                  cmp = request.user.company
                else:
                  cmp = request.user.employee.company  
                usr = CustomUser.objects.get(username=request.user)
                # print(emails_list)
                pbill = PurchaseBill.objects.get(id=id,company=cmp)
                pitm = PurchaseBillItem.objects.filter(purchasebill=pbill)
                dis = 0
                for itm in pitm:
                  dis += int(itm.discount)
                itm_len = len(pitm)
                context={'pbill':pbill,'pitm':pitm,'itm_len':itm_len,'dis':dis}
                template_path = 'vatpdf.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                filename = f'Purchase bill - {pbill.billno}.pdf'
                subject = f"Purchase bill - {pbill.billno}"
                email = EmailMessage(subject, f"Hi,\nPlease find the Purchase bill- Bill-{pbill.billno}. \n{email_message}\n\n--\nRegards,\n{pbill.company.company_name}\n{pbill.company.address}\n - {pbill.company.city}\n{pbill.company.contact}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'purchase bill has been shared via email successfully..!')
                return redirect(details_purchasebill,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(details_purchasebill, id)
            
def save_party2(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    usr = CustomUser.objects.get(username=request.user)

    if request.method == 'POST':
        pbillhide=request.POST.get('pbillhid')
        partyname = request.POST.get('partyname')
        trn_no = request.POST.get('trn_no')
        contact = request.POST.get('contact')
        trn_type = request.POST.get('trn_type')
        address = request.POST.get('address')
        email = request.POST.get('email')
        balance = request.POST.get('balance')
        paymentType = request.POST.get('paymentType')
        currentdate = request.POST.get('currentdate')
        additionalfield1 = request.POST.get('additionalfield1')
        additionalfield2 = request.POST.get('additionalfield2')
        additionalfield3 = request.POST.get('additionalfield3')
        # print(trn_no)
        # print(partyname)

        # Check if the contact number already exists in the database
        if Party.objects.filter(contact=contact,company=cmp).exists():
            # Send a message indicating that the contact number already exists

            return redirect('edit_purchasebill', pbillhide)

        # Check if the transaction number already exists in the database
        if trn_type == "Unregistered/Consumers":
            prtobj = Party.objects.create(    user=usr,
            company=cmp,
            party_name=partyname,
            trn_no=trn_no,
            contact=contact,
            trn_type=trn_type,
            address=address,
            email=email,
            openingbalance=balance,
            payment=paymentType,
            current_date=currentdate,
            additionalfield1=additionalfield1,
            additionalfield2=additionalfield2,
            additionalfield3=additionalfield3)
            prtobj.save()
            # Optionally, you can send a success message here

            return redirect('edit_purchasebill', pbillhide)
        else:
             if Party.objects.filter(trn_no=trn_no, company=cmp).exists():
                # Send a message indicating that the transaction number already exists

                return redirect('edit_purchasebill', pbillhide)
        prtobj = Party.objects.create(
            user=usr,
            company=cmp,
            party_name=partyname,
            trn_no=trn_no,
            contact=contact,
            trn_type=trn_type,
            address=address,
            email=email,
            openingbalance=balance,
            payment=paymentType,
            current_date=currentdate,
            additionalfield1=additionalfield1,
            additionalfield2=additionalfield2,
            additionalfield3=additionalfield3
        )
        prtobj.save()
        return redirect('edit_purchasebill', pbillhide)
        
def check_trn_no_exists(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    usr = CustomUser.objects.get(username=request.user) 
    trn_no = request.GET.get('trn_no')
    trn_type = request.POST.get('trn_type')
    print(trn_type)
    if trn_type != "Unregistered/Consumers":
      if Party.objects.filter(trn_no=trn_no,company=cmp).exists():
          return JsonResponse({'exists': True})
      return JsonResponse({'exists': False})

def check_phone_number_exists(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    usr = CustomUser.objects.get(username=request.user)
    phone_number = request.GET.get('phone_number')
    if Party.objects.filter(contact=phone_number,company=cmp).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists': False})
    
def check_hsn_number_exists(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    usr = CustomUser.objects.get(username=request.user)
    hsn = request.GET.get('hsn1')
    if Item.objects.filter(itm_hsn=hsn,company=cmp).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists': False})
    
    
def unit_reload_modal(request):
    if request.user.is_authenticated:
        if request.user.is_company:
            cmp = request.user.company
        else :
            cmp = request.user.employee.company
        units = Unit.objects.filter(company=cmp)
        
        units_list = [{'name': unlist.unit_name} for unlist in units]
        success = True
        data = {'unitsobj': units_list,'success': success}
        return JsonResponse(data,safe=False)
       
    else:
        # Handle the case when the user is not authenticated
        return JsonResponse({'error': 'User is not authenticated'},status=401)
        
        
#Aami Jafer MultiUser Billing  softwareVAT
def creditNote(request):
  if request.user.is_company:
        cmp = request.user.company
  else:
        cmp = request.user.employee.company
  creditnotes=CreditNote.objects.filter(company=cmp)
  if not creditnotes:
    if CreditNoteReference.objects.filter(company=cmp).exists():
      return redirect('SalesReturn') 
    else:
      return render(request,'CreditNote.html',{'usr':request.user})
  else:
    return redirect('SalesReturn') 

def SalesReturn(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company
    parties = Party.objects.filter(company=cmp)
    items = Item.objects.filter(company=cmp)
    unit = Unit.objects.filter(company=cmp)
    
    try:
      max_reference_no = CreditNote.objects.filter(company=cmp).last().reference_no
    except AttributeError:
      max_reference_no = None

    if max_reference_no is not None:
        reference_no = max_reference_no + 1
    else:
        reference_no = 1
# Check if there are deleted credit notes in CreditNoteReference
  
    context = {'usr':request.user, 'parties':parties, 'items':items,'unit':unit,'cmp':cmp,'reference_number': reference_no}
    
    return render(request, 'SalesReturn.html', context)

def saveParty(request):
    if request.user.is_authenticated:
      if request.method == "POST":
          if request.user.is_company:
            cmp = request.user.company
          else:
            cmp = request.user.employee.company  
          print(cmp)
          usr = CustomUser.objects.get(username=request.user)
          party_name = request.POST['party_name']
          gst_no = request.POST['gst_no']
          mob = request.POST['mob']
          gsttype = request.POST['gsttype']
          state = request.POST['state']
          email = request.POST['email']
          addr = request.POST['addr']
          opbal = request.POST['opbal']
          payment = request.POST.get('paymentType','')
          date = request.POST['date']
          add1 = request.POST['add1']
          add2 = request.POST['add2']
          add3 = request.POST['add3']
          if Party.objects.filter(trn_no=gst_no, company=cmp).exists() and not(gsttype == 'Unregistered/Consumers'):
            error_message = 'TRN number already exists!!!'
            return JsonResponse({'error_message': error_message})
          if gst_no == '' and (gsttype == 'Registered Business - Regular' or gsttype == 'Registered Business - Composition'):
            error_message = 'TRN number required.'
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(email=email, company=cmp).exists():
            error_message = 'Email already exists!!!'
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(contact=mob,company=cmp).exists():
            error_message = 'Mobile number already exists!!!'
            return JsonResponse({'error_message': error_message})
          user = request.user  
          party = Party(
                  user=user,
                  company=cmp,
                  party_name=party_name,
                  trn_no=gst_no,
                  contact=mob,
                  trn_type=gsttype,
                  state=state,
                  address=addr,
                  email=email,
                  openingbalance=opbal,
                  payment=payment,
                  current_date=date,
                  additionalfield1=add1,
                  additionalfield2=add2,
                  additionalfield3=add3
              )
          party.save()
          print('Party created succefully ')
          trans = Transactions_party(user=request.user,company=cmp, trans_type='Opening Balance', trans_number=gst_no,
                                           trans_date=date, total=opbal, balance=opbal, party=party)
          trans.save()
          tr_history = PartyTransactionHistory(party=party, Transactions_party=trans, action="CREATED")
          tr_history.save()
          return JsonResponse({'success': True})
      
def party_dropdown(request):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company  
  options={}
  option_objects=Party.objects.filter(company=cmp)
  for option in option_objects:
    options[option.id]=option.party_name
  return JsonResponse(options)

def get_partydetails(request):
    if request.method == 'POST':
        if request.user.is_company:
            cmp = request.user.company
        else:
            cmp = request.user.employee.company
            
        party_id = request.POST.get('id').split(" ")[0]
        party = get_object_or_404(Party, company=cmp, id=party_id)
        print(party_id)

        balance = party.openingbalance if party.openingbalance else None
        phone = party.contact if party.contact else None
        address = party.address if party.address else None
        payment1 = party.payment if party.payment else None

        sales_invoice =SalesInvoice.objects.filter(company=cmp, party=party).first()
        invoiceno = sales_invoice.invoice_no if sales_invoice else None
        invoicedate = sales_invoice.invoice_date if sales_invoice else None
        placeofsupply = sales_invoice.address if sales_invoice else None
        payment = sales_invoice.payment if sales_invoice else None

        return JsonResponse({
            'address': address,
            'bal': balance,
            'phone': phone,
            'invoiceno': invoiceno,
            'invoicedate': invoicedate,
            'placeofsupply': placeofsupply,
            'pay': payment,
            'payment1':payment1,
            'id':party.id,
        })

from django.views.decorators.http import require_http_methods
@require_http_methods(["GET", "POST"])
def get_invoice_date(request):
    if request.method == 'POST':
        invoiceno = request.POST.get('invoiceno', None)
    elif request.method == 'GET':
        invoiceno = request.GET.get('invoiceno', None)
    
    print('invoiceno:', invoiceno)

    if not invoiceno:
        #print('5')
        return JsonResponse({'error': 'Invoice number not provided'}, status=400)  # Bad Request

    try:
        # Get the latest SalesInvoice with the specified invoice_no
        invoice = Invoice.objects.get(invoice_no=invoiceno)
        print('invoice1:', invoice)
        invoicedate = invoice.invoice_date.strftime('%Y-%m-%d')
        print('invoicedate:', invoicedate)
        print('1:',invoice.invoice_date)
    except Invoice.DoesNotExist:
        #print('3')
        return JsonResponse({'error': 'Invoice number not found'}, status=404)  # Not Found
    except Invoice.MultipleObjectsReturned:
        #print('2')
        return JsonResponse({'error': 'Multiple invoices found for the same invoice number'}, status=400)  # Bad Request
    except Exception as e:
        #print('1')
        return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)  # Internal Server Error
    #print('4')
    return JsonResponse({'invoicedate':invoicedate})

def get_invoice_item(request):

    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    invoiceno = request.GET.get('invoiceno') 
    print(invoiceno, 'ftydf')  # Output the invoice number for debugging
    try:
        # Retrieve the invoice object with the given invoice number or return a 404 error if not found
        invoice = get_object_or_404(SalesInvoice, invoice_no=invoiceno)
        invoice_items = SalesInvoiceItem.objects.filter(company=cmp, salesinvoice=invoice)
    except SalesInvoice.DoesNotExist:
        return redirect(SalesReturn)
    return render(request, 'tab_logic.html',{"items":invoice_items})
    
def saveItem(request):
  if request.method == 'POST':
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company  
    print(cmp)
    usr = CustomUser.objects.get(username=request.user)
    item_type=request.POST['item_type']
    item_name=request.POST['item_name']
    item_hsn=request.POST['item_hsn']
    item_unit=request.POST['item_unit']
    itm_taxable=request.POST['itm_taxable']
    itm_vat=request.POST['itm_vat']
    itm_sale_price=request.POST['itm_sale_price']
    itm_purchase_price=request.POST['itm_purchase_price']
    itm_stock_in_hand=request.POST['itm_stock_in_hand']
    itm_at_price=request.POST['itm_at_price']
    itm_date=request.POST['itm_date']
    if Item.objects.filter(company=cmp,itm_name=item_name).exists():
      error_message = 'Item already registered!!!'
      return JsonResponse({'error_message': error_message})
    if Item.objects.filter(company=cmp,itm_hsn=item_hsn).exists():
      error_message = 'Item with this HSN already registered!!!'
      return JsonResponse({'error_message': error_message})
    item = Item(
                user=usr,
                company=cmp,
                itm_type=item_type,
                itm_name=item_name,
                itm_hsn=item_hsn,
                itm_unit=item_unit,
                itm_taxable=itm_taxable,
                itm_vat=itm_vat,
                itm_sale_price=itm_sale_price,
                itm_purchase_price=itm_purchase_price,
                itm_stock_in_hand=itm_stock_in_hand,
                itm_at_price=itm_at_price,
                itm_date=itm_date
            )
    print(item)
    item.save()
    trans = ItemTransactions(user = request.user, item = item, trans_type = 'Stock Open', trans_date = itm_date, trans_qty = itm_stock_in_hand, 
                             trans_current_qty = itm_stock_in_hand, trans_adjusted_qty = itm_stock_in_hand, trans_price = itm_at_price)
    trans.save()
    trhis = ItemTransactionsHistory(user = request.user, transaction = trans, hist_trans_qty = 0, hist_trans_current_qty = itm_stock_in_hand, action = 'Created',
                                    hist_trans_adjusted_qty = itm_stock_in_hand, hist_trans_type = trans.trans_type)
    trhis.save()
    print("Item saved")
    return HttpResponse({"message": "success"})
  
    
def get_itemdetails(request):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company  
  item_id = request.POST.get('id').split(" ")[0]
  item=Item.objects.get(company=cmp,pk=item_id)
  hsn=item.itm_hsn
  price=item.itm_sale_price
  tax=item.itm_vat
  return JsonResponse({'hsn':hsn,'price':price,'tax':tax})

def fetch_item_details(request):
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company
    item_id = request.POST.get('id')
    item = Item.objects.get(pk=item_id,company=cmp)
    data = {
                'hsn': item.itm_hsn,
                'price': item.itm_sale_price,
                'tax': item.itm_vat
            }
    return JsonResponse(data)

def get_item_dropdown(request):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company  
  options={}
  option_objects=Item.objects.filter(company=cmp)
  for option in option_objects:
    options[option.id]=[option.id, option.itm_name, option.itm_hsn, option.itm_sale_price,option.itm_vat]
  return JsonResponse(options)

def saveCreditnote(request):
  if request.method == 'POST':
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company
    usr = CustomUser.objects.get(username=request.user)
    print("usr:",usr)
    return_date=request.POST.get('returndate')
    print("returndate:",return_date)
    reference_no=request.POST.get('refnum')
    print("reference_no", reference_no)
    subtotal=request.POST.get('sub_total')
    print("sub: ",subtotal)
    vat=request.POST.get('disvatper')
    print("VAT: ",vat)
    adjustment=request.POST.get('adjustment')
    print("adj:",adjustment)
    grandtotal=request.POST.get('grandtotal')
    print("grandtotal: ",grandtotal)
    party_status = request.POST.get('partystatus')
    print("Partystatus: ",party_status)
    
    creditnote_curr = CreditNote(
      user=usr, 
      company=cmp,
      reference_no=reference_no,
      partystatus=party_status,
      returndate=return_date,
      subtotal=subtotal,
      vat=vat,
      adjustment=adjustment,
      grandtotal=grandtotal
      )
    creditnote_curr.save()
    print("creditnote_curr:",creditnote_curr)
    if party_status=='partyon':
      party_details = request.POST.get('party_details')
      print("party_details:",party_details)
      party_id = party_details.split()[0]
      print("party_id:",party_id)
      party = Party.objects.get(pk=party_id,company=cmp)
      
      print(party.party_name)
      creditnote_curr.party=party
      creditnote_curr.save()
      print("creditnote_curr.party:",creditnote_curr.party)
      
      salesinvoice=request.POST.get('invoiceno')
      print("salesinvoice:",salesinvoice)
      idsales=salesinvoice.split()[0]
      print("idsales:",idsales)
      invoice=Invoice.objects.get(invoice_no=idsales,company=cmp)
      print(invoice.invoice_no)
      creditnote_curr.salesinvoice=invoice
      creditnote_curr.save()
      # salesinvoice =Invoice.objects.filter(company=cmp, party=party)
      # for invoice in salesinvoice:
      #     invoiceno.append(invoice.invoice_no)
      # if salesinvoice:
      #   idsales=request.POST['invoiceno']
      #   creditnote_curr.salesinvoice=Invoice.objects.get(invoice_no=idsales,company=cmp)
      #   creditnote_curr.save()
      # else:
      #   pass
    item_name =request.POST.getlist('item_name[]')
    print("item name: ",item_name)
    quantity =request.POST.getlist('qty[]')
    print("item qty: ",quantity)
    price =request.POST.getlist('price[]')
    print("item price",price)
    tax =request.POST.getlist('tax[]')
    print("item tax",tax)
    discount = request.POST.getlist('discount[]')
    print("item discount",discount)
    hsn = request.POST.getlist('hsn[]')
    print("item hsn",hsn)
    total = request.POST.getlist('total[]')
    print("item total",total)
    
    if len(quantity) == len(price) == len(tax) == len(discount) == len(hsn) == len(total) and item_name and quantity and price and tax and discount and hsn and total:
      mapped=zip(item_name,quantity,price,tax,discount,hsn,total)
      mapped=list(mapped)
      for ele in mapped:
        print("Element:", ele)
        hsn = ele[5]
        quantity = ele[1]
        tax = ele[3]
        price = ele[2]
        discount = ele[4]
        total = ele[6]
        print("HSN:", hsn)
        print("Quantity:", quantity)
        print("Tax:", tax)
        print("Price:", price)
        print("Discount:", discount)
        print("Total:", total)

        item_name_parts = ele[0].split()
        print("item_name_parts: ",item_name_parts)
        item_id = item_name_parts[0]
        print("item_id: ",item_id)
        items = Item.objects.get(id=item_id,company=cmp)
        print("items are: ",items)
        it=Item.objects.get(company=cmp, id = item_id).itm_name
        print("item_name:", it)
        # creditnote = CreditNote.objects.get(reference_no=creditnote_curr.reference_no,company=cmp)
        # print('1:',creditnote)
        creditnoteitem=CreditNoteItem(
                  user=usr,
                  credit_note=creditnote_curr,
                  company=cmp,
                  items=items,
                  item=it,
                  hsn=ele[5],
                  quantity=ele[1],
                  tax=ele[3],
                  price=ele[2],
                  discount=ele[4],
                  total=ele[6])
        creditnoteitem.save()
        print("creditnoteitem:", creditnoteitem)
      
      itt = ItemTransactions.objects.filter(item=items.id).last().trans_current_qty

      # Add invoice details in items transactions
      transaction = ItemTransactions.objects.create(
          company=cmp,
          item=items,
          trans_type='Creditnote',
          trans_date=creditnote_curr.creditnote_date,
          trans_qty=ele[1],
          trans_current_qty = itt + int(ele[1]),
          trans_price=ele[2],
          #trans_invoice=invoice_curr.invoice_no
      )
      items.itm_stock_in_hand-= int(ele[1])
      items.save()
      #transaction.trans_current_qty -= int(ele[1]) 
      #transaction.save()
      print("Transaction saved:", transaction)
    
    p_id  = Party.objects.get(id=request.POST.get('partyname'))
    print('transaction0')
    p_id.openingbalance=float(p_id.openingbalance)-float(creditnote_curr.grandtotal)
    p_id.save()
    transactionparty=Transactions_party.objects.create(
      company=cmp,
      user=usr,
      party=p_id,
      trans_type='Creditnote',
      trans_number=creditnote_curr.reference_no,
      trans_date=creditnote_curr.creditnote_date,
      total=creditnote_curr.grandtotal,
      balance=p_id.openingbalance,
      
    )
    print(transactionparty.party)   
    print("Transaction party1:", transactionparty)
    transactionparty.save()


    creditnote = CreditNote.objects.get(reference_no=creditnote_curr.reference_no,company=cmp)
    print('1:',creditnote)
    cr=CreditNoteHistory(user=usr,company=cmp,credit_note_history=creditnote,action='Created')
    cr.save()
    if 'save_new' in request.POST:
      return redirect('SalesReturn')
    else:
      return redirect('listout_page') 
  
def listout_page(request):
  if request.user.is_company:
          cmp = request.user.company
  else:
          cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
   
  itm=CreditNote.objects.filter(company=cmp)
  crbill =CreditNote.objects.filter(company=cmp).values()
  crbills =CreditNote.objects.filter(company=cmp)

  for i in crbill:
    p_history = CreditNoteHistory.objects.filter(credit_note_history=i['id'], company=cmp).last()
    if p_history:
        i['action'] = p_history.action
        i['name']= f"{p_history.user.first_name} {p_history.user.last_name}"
        i['party'] = p_history.credit_note_history.party
    else:
        # Handle the case when no history is found
        i['action'] = ""
        i['name'] = ""
        i['party_name'] = ""


  return render(request, 'listout.html',{'itm':itm,'crbill':crbill,'crbills':crbills,'usr':request.user})
  
def edit_creditnote(request,pk):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company
  parties = Party.objects.filter(company=cmp)
  items = Item.objects.filter(company=cmp)
  unit = Unit.objects.filter(company=cmp)
  creditnote_curr=CreditNote.objects.get(id=pk,company=cmp)
  print('1:',creditnote_curr.salesinvoice.invoice_no)
  print('2:',creditnote_curr.salesinvoice.invoice_date)
  name=creditnote_curr.party
  reference=creditnote_curr.reference_no
  print(name)

  creditnote_items=CreditNoteItem.objects.filter(credit_note=creditnote_curr,company=cmp)
  for item in creditnote_items:
    print(f"Item ID: {item.id}")
    print(f"Credit Note: {item.item}")
  context={'usr':request.user,
           'creditnoteitem_curr':creditnote_items,
           'credit_note':creditnote_curr,
           'parties':parties,
           'reference':reference,
           'items':items,'unit':unit,
          
          }
  return render(request,'edit_creditnote.html',context)

def delete_creditnote(request,pk):
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  creditnote=CreditNote.objects.get(id=pk,company=cmp)
  reference_no=creditnote.reference_no
  CreditNoteItem.objects.filter(credit_note= creditnote,company=cmp).delete()
  cr=CreditNoteReference.objects.create(user=request.user,company=cmp,reference_no=reference_no)
  cr.save()
  creditnote.delete()
  return redirect('listout_page')

def updateCreditnote(request, pk):
    if request.method == "POST":
        if request.user.is_company:
            cmp = request.user.company
            print('cmp1:',cmp)
        else:
            cmp = request.user.employee.company
            print('cmp2:',cmp)

        usr = get_object_or_404(CustomUser, username=request.user)
        print('usr:',usr)
        creditnote = get_object_or_404(CreditNote, id=pk, company=cmp)
        print('creditnote:',creditnote)
        creditnote.user = request.user
        print('creditnote.user:',creditnote.user)
        creditnote.company = cmp
        print('creditnote.company:',creditnote.company)
        creditnote.returndate = request.POST.get('returndate')
        print('creditnote.returndate:',creditnote.returndate)
        creditnote.reference_no = request.POST.get('refnum')
        print('creditnote.reference_no:',creditnote.reference_no)
        creditnote.subtotal = request.POST.get('subtotal')
        print('creditnote.subtotal:',creditnote.subtotal)
        creditnote.vat = request.POST.get('disvatper')
        print('creditnote.vat:',creditnote.vat)
        creditnote.adjustment = request.POST.get('adjustment')
        print('creditnote.adjustment:',creditnote.adjustment)
        creditnote.grandtotal = request.POST.get('grandTotal')
        print('creditnote.grandtotal:',creditnote.grandtotal)
        creditnote.partystatus = request.POST.get('partystatus')
        print('creditnote.partystatus:',creditnote.partystatus)
        creditnote.save()
        print('creditnote:',creditnote)

        if creditnote.partystatus == 'partyon':
            party_details = request.POST.get('party_details')
            print('party_details:',party_details)
            party_id = party_details.split()[0]
            print('party_id:',party_id)
            party = get_object_or_404(Party, pk=party_id, company=cmp)
            print('party :',party)
            party.contact = request.POST.get('party_phone')
            print('party.contact:',party.contact)
            party.payment = request.POST.get('paymethod')
            print('party.payment:',party.payment)
            party.save()
            print('party:',party)
            creditnote.party = party
            print('creditnote.party:',creditnote.party)

            idsales = request.POST.get('invoiceno')
            print('idsales:',idsales)
            salesinvoice =Invoice.objects.filter(company=cmp, party=party, invoice_no=idsales).first()
            print('salesinvoice:',salesinvoice)
            if salesinvoice:
                creditnote.salesinvoice = salesinvoice
                print('creditnote.salesinvoice:',creditnote.salesinvoice)
            creditnote.save()
            print('creditnote:',creditnote)
        else:
            creditnote.party = None
            creditnote.save()
            print('creditnote:',creditnote)

        CreditNoteHistory.objects.create(user=usr, company=cmp, credit_note_history=creditnote, action='Updated')
        

        item_name = request.POST.getlist('item_name[]')
        print('item_name:',item_name)
        quantity = request.POST.getlist('qty[]')
        print('quantity:',quantity)
        price = request.POST.getlist('price[]')
        print('price:',price)
        tax = request.POST.getlist('tax[]')
        print('tax:',tax)
        discount = request.POST.getlist('discount[]')
        print('discount:',discount)
        hsn = request.POST.getlist('hsn[]')
        print('hsn:',hsn)
        total = request.POST.getlist('total[]')
        print('total:',total)

        if len(item_name) == len(quantity) == len(price) == len(tax) == len(discount) == len(hsn) == len(total):
            mapped = zip(item_name, quantity, price, tax, discount, hsn, total)
            credit_note_items = CreditNoteItem.objects.filter(credit_note=creditnote, company=cmp)
            print('credit_note_items',credit_note_items)
            existing_item_ids = []
            print('existing_item_ids',existing_item_ids)

            for ele in mapped:
                item_id = ele[0].split()[0]
                print('item_id',item_id)
                #items = get_object_or_404(Item, company=cmp, pk=item_id)
                items = Item.objects.get(pk=item_id,company=cmp)
                print('items',items)
                existing_credit_note_item = credit_note_items.filter(items=items).first()
                print('existing_credit_note_item',existing_credit_note_item)

                if existing_credit_note_item:
                    existing_credit_note_item.quantity = ele[1]
                    print('existing_credit_note_item.quantity',existing_credit_note_item.quantity)
                    existing_credit_note_item.price = ele[2]
                    print('existing_credit_note_item.price',existing_credit_note_item.price)
                    existing_credit_note_item.tax = ele[3]
                    print('existing_credit_note_item.tax',existing_credit_note_item.tax)
                    existing_credit_note_item.discount = ele[4]
                    print('existing_credit_note_item.discount',existing_credit_note_item.discount)
                    existing_credit_note_item.hsn = ele[5]
                    print('existing_credit_note_item.hsn',existing_credit_note_item.hsn)
                    existing_credit_note_item.total = ele[6]
                    print('existing_credit_note_item.total',existing_credit_note_item.total)
                    existing_credit_note_item.save()
                    print('existing_credit_note_item',existing_credit_note_item)
                else:
                    CreditNoteItem.objects.create(
                        user=usr,
                        credit_note=creditnote,
                        company=cmp,
                        items=items,
                        item=items.itm_name,
                        hsn=ele[5],
                        quantity=ele[1],
                        tax=ele[3],
                        price=ele[2],
                        discount=ele[4],
                        total=ele[6]
                    )
                existing_item_ids.append(items.id)

            items_to_delete = credit_note_items.exclude(items__id__in=existing_item_ids)
            print('items_to_delete',items_to_delete)
            items_to_delete.delete()

            return redirect('listout_page')

        return HttpResponse('Invalid input data', status=400)

    return HttpResponse('Invalid request method', status=405)
    
def credit_templates(request,pk):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company 
  creditnote_curr=CreditNote.objects.get(id=pk,company=cmp)
  reference=creditnote_curr.reference_no
  creditnote_items=CreditNoteItem.objects.filter(credit_note=creditnote_curr,company=cmp)
  context = {'usr':request.user,'company':cmp,'creditnoteitem_curr':creditnote_items,
           'creditnote':creditnote_curr,'reference':reference}
  return render(request,'creditnote_temp.html',context)

def sharebill(request,id):
 if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                if request.user.is_company:
                  cmp = request.user.company
                else:
                  cmp = request.user.employee.company  
                usr = CustomUser.objects.get(username=request.user)
                # print(emails_list)
                creditnote = CreditNote.objects.get(id=id,company=cmp)
                creditnoteitem_curr = CreditNoteItem.objects.filter(credit_note=creditnote)
                dis = 0
                for itm in creditnoteitem_curr:
                  dis += int(itm.discount)
                itm_len = len(creditnoteitem_curr)
                context={'creditnote':creditnote,'creditnoteitem_curr':creditnoteitem_curr,'itm_len':itm_len,'dis':dis}
                template_path = 'creditnotepdf.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                filename = f'Creditnote - {creditnote.reference_no}.pdf'
                subject = f"Creditnote - {creditnote.reference_no}"
                email = EmailMessage(subject, f"Hi,\nPlease find the Invoice bill- Bill-{creditnote.reference_no}. \n{email_message}\n\n--\nRegards,\n{creditnote.company.company_name}\n{creditnote.company.address}\n - {creditnote.company.city}\n{creditnote.company.contact}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Creditnote has been shared via email successfully..!')
                return redirect(credit_templates,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(credit_templates, id) 

def render_to_pdf(html):
    # Generate PDF from HTML content
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return result.getvalue()
    return None

def history_page(request,pk):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company
  creditnote=CreditNote.objects.get(id=pk,company=cmp)
  reference=creditnote.reference_no
  print(reference)
  credit_hist=CreditNoteHistory.objects.filter(credit_note_history=pk,company=cmp) 
  context={'c_usr':request.user,'c_comp':cmp,'creditnote':creditnote,'credit_hist':credit_hist,'reference':reference}
  return render(request,'historyPage.html',context)

#Aami Jafer MultiUser Billing  softwareVAT

def party_list(request):
  if request.user.is_company:
    party = Party.objects.filter(company = request.user.company)
  else:
    party = Party.objects.filter(company = request.user.employee.company)
  
  if party:
    fparty = party[0]
    ftrans = Transactions_party.objects.filter(party = fparty)
    balance = Transactions_party.objects.filter(party = fparty).last()
    context = {'party':party, 'usr':request.user, 'fparty':fparty, 'ftrans':ftrans,'balance':balance}
  else:
        context = {'party':party, 'usr':request.user}
  return render(request,'parties_list.html',context)


def load_party_create(request):
  tod = timezone.now().date().strftime("%Y-%m-%d")
  return render(request,'add_parties.html',{'tod':tod, 'usr':request.user})


def addNewParty(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company

    error_message = None  # Initialize error message

    if request.method == 'POST':
        party_name = request.POST['partyname']
        trn_no = request.POST['trn_no']
        contact = request.POST['contact']
        trn_type = request.POST['trn_type']
        state = request.POST.get('state')
        address = request.POST['address']
        email = request.POST['email']
        opening_stock=request.POST['opening_stock']
        at_price=request.POST['at_price']
        openingbalance = request.POST.get('balance', '')
        payment = request.POST.get('paymentType', '')
        current_date = request.POST['currentdate']
        End_date = request.POST.get('enddate', None)
        additionalfield1 = request.POST['additionalfield1']
        additionalfield2 = request.POST['additionalfield2']
        additionalfield3 = request.POST['additionalfield3']

        try:
            # Check if a party with the same trn_no already exists
            if Party.objects.filter(trn_no=trn_no, company=cmp).exists():
              error_message = 'An error occurred while processing your request. TRN number already exists. Please enter a unique TRN number.'
              print("error_message1",error_message)
            # Check if a party with the same email already exists
            elif Party.objects.filter(email=email, company=cmp).exists():
              error_message = 'An error occurred while processing your request. Email already exists. Please enter a unique email address.'
              print("error_message2",error_message)
            
            elif Party.objects.filter(contact=contact,company=cmp).exists():
              error_message = 'An error occurred while processing your request. Mobile number already exists. Please enter a unique mobile number.'
              print("error_message3",error_message)
              #return JsonResponse({'error_message': error_message})
            
            else:
                part = Party(party_name=party_name, trn_no=trn_no, contact=contact, trn_type=trn_type, state=state,
                             address=address, email=email, openingbalance=openingbalance,opening_stock=opening_stock,at_price=at_price ,payment=payment,
                             current_date=current_date, End_date=End_date, additionalfield1=additionalfield1,
                             additionalfield2=additionalfield2, additionalfield3=additionalfield3)

                if request.user.is_company:
                    part.company = request.user.company
                else:
                    part.company = request.user.employee.company

                part.save()

                trans = Transactions_party(user=request.user, trans_type='Opening Balance', trans_number=trn_no,
                                           trans_date=current_date, total=openingbalance, balance=openingbalance, party=part)

                if request.user.is_company:
                    trans.company = request.user.company
                else:
                    trans.company = request.user.employee.company

                trans.save()

                tr_history = PartyTransactionHistory(party=part, Transactions_party=trans, action="CREATED")
                tr_history.save()

                if request.POST.get('save_and_next'):
                    return redirect('load_party_create')
                elif request.POST.get('save'):
                    return redirect('party_list')

        except IntegrityError:
            # Specific error message for duplicate TRN number
            error_message = 'An error occurred while processing your request. Please try again.'

    return render(request, 'add_parties.html', {'error_message': error_message})


def view_party(request,id):
  if request.user.is_company:
    party = Party.objects.filter(company = request.user.company)
  else:
    party = Party.objects.filter(company = request.user.employee.company)

  fparty = Party.objects.get(id=id)
  ftrans = Transactions_party.objects.filter(party = fparty)

  total_trans = Transactions_party.objects.filter(party = fparty).aggregate(total_trans=Sum('balance'))['total_trans'] or 0
  
  context = {'party':party, 'usr':request.user, 'fparty':fparty, 'ftrans':ftrans,'total_trans':total_trans}
  return render(request,'parties_list.html',context)


def edit_party(request,id):
  if request.user.is_company:
    party = Party.objects.filter(company = request.user.company)
  else:
    party = Party.objects.filter(company = request.user.employee.company)

  getparty=Party.objects.get(id=id)
  parties=Party.objects.filter(user=request.user)
  ftrans = Transactions_party.objects.filter(party = getparty)
  return render(request, 'edit_party.html',{'usr':request.user,'party':party,'getparty':getparty,'parties':parties,'ftrans':ftrans})


def edit_saveparty(request, id):
    company = request.user.company if request.user.is_company else request.user.employee.company
    party_qs = Party.objects.filter(company=company)
    getparty = get_object_or_404(Party, id=id, company=company)
    
    trans = None  # Initialize trans here

    if request.method == 'POST':
        # Update party details
        getparty.party_name = request.POST.get('partyname')
        getparty.trn_no = request.POST.get('trn_no')
        getparty.contact = request.POST['contact']
        getparty.trn_type = request.POST['trn_type']
        getparty.state = request.POST['state']
        getparty.address = request.POST['address']
        getparty.email = request.POST['email']
        getparty.openingbalance = request.POST['balance']
        getparty.payment = request.POST.get('paymentType')
        # getparty.current_date = request.POST['currentdate']
        getparty.current_date = datetime.now().strftime('%Y-%m-%d')
        getparty.additionalfield1 = request.POST['additionalfield1']
        getparty.additionalfield2 = request.POST['additionalfield2']
        getparty.additionalfield3 = request.POST['additionalfield3']

        # Save the party changes
        getparty.save()

        # Check if a transaction with the same party_id exists
        existing_transaction = Transactions_party.objects.filter(party_id=id).first()

        # Update or create a new transaction
        if existing_transaction:
            existing_transaction.trans_type = 'Opening Balance'
            existing_transaction.trans_number = getparty.trn_no
            existing_transaction.trans_date = getparty.current_date
            existing_transaction.total = getparty.openingbalance
            existing_transaction.balance = getparty.openingbalance
            existing_transaction.save()

            # Check and update the transaction history action to "UPDATED"
            party_history_entry = PartyTransactionHistory.objects.filter(
                party=getparty, Transactions_party=existing_transaction
            ).first()
            if party_history_entry:
                party_history_entry.action = "UPDATED"
                party_history_entry.save()

        else:
            # Create and save a new transaction
            trans = Transactions_party(
                user=request.user,
                trans_type='Opening Balance',
                trans_number=getparty.trn_no,
                trans_date=getparty.current_date,
                total=getparty.openingbalance,
                balance=getparty.openingbalance,
                party=getparty,
                company=company
            )
            trans.save()

            # Save the transaction history entry with "CREATED" action
            tr_history = PartyTransactionHistory(
                party=getparty,
                Transactions_party=trans,
                action="UPDATED"
            )
            tr_history.save()

        return redirect('view_party', id=getparty.id)

    return render(request, 'edit_party.html', {'getparty': getparty, 'party': party_qs, 'usr': request.user})


def history_party(request, id):
    if request.user.is_company:
        party = Party.objects.filter(company=request.user.company)
    else:
        party = Party.objects.filter(company=request.user.employee.company)

    fparty = Party.objects.get(id=id)
    #ftrans = Transactions_party.objects.get(party=fparty)
    hst = PartyTransactionHistory.objects.filter(party=id)

    context = {'party': party, 'hst': hst, 'usr': request.user, 'fparty': fparty}
    return render(request, 'partyhistory.html',context)


def deleteparty(request,id):
    if request.user.is_company:
      party = Party.objects.filter(company = request.user.company)
    else:
      party = Party.objects.filter(company = request.user.employee.company)

    party=Party.objects.get(id=id)
    party.delete()
    return redirect('party_list')
    

def shareTransactionpartyToEmail(request, id):
    if request.user.is_company:
        party = Party.objects.filter(company=request.user.company)
    else:
        party = Party.objects.filter(company=request.user.employee.company)
        
    if request.method == "POST":
        try:
         
            fparty = Party.objects.get(id=id)
            ftrans = Transactions_party.objects.filter(party=fparty)
            cmp = fparty.company
            context = {'party': party, 'usr': request.user, 'fparty': fparty, 'ftrans': ftrans,'cmp':cmp}
            
            email_message = request.POST.get('email_message')
            my_subject = "Transaction REPORT"
            emails_string = request.POST.get('email_ids')
            emails_list = [email.strip() for email in emails_string.split(',')]
            
            html_message = render_to_string('transaction_pdf.html', context)
            plain_message = strip_tags(html_message)
            
            pdf_content = BytesIO()
            pisa.CreatePDF(html_message.encode("UTF-8"), pdf_content)
            pdf_content.seek(0)

            filename = f'transaction.pdf'
            email=EmailMultiAlternatives(my_subject,f"Hi,\nPlease find the attached Transaction Report - \n{email_message}\n--\nRegards,\n{cmp.company_name},\n{cmp.address},{cmp.city},{cmp.country},\n{cmp.contact}\n",from_email='altostechnologies6@gmail.com',
                                         to=emails_list, )
       
            email.attach(filename, pdf_content.read(), 'application/pdf')
            email.send()

            return HttpResponse('<script>alert("Report has been shared successfully!");window.location="/party_list"</script>')
        except Party.DoesNotExist:
            return HttpResponse('<script>alert("Party not found!");window.location="/party_list"</script>')
        except Exception as e:
            # Handle the exception, log the error, or provide an error message
            return HttpResponse(f'<script>alert("Failed to send email: {str(e)}");window.location="/party_list"</script>')

    return HttpResponse('<script>alert("Invalid Request!");window.location="/party_list"</script>')
    
    
def firstdebitnote(request):
  if request.user.is_company:
          cmp = request.user.company
  else:
          cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
   
  itm=DebitNote.objects.filter(company=cmp)
  dbill = DebitNote.objects.filter(company=cmp).values()
  dbills = DebitNote.objects.filter(company=cmp)

  for i in dbill:
    p_history = DebitNoteHistory.objects.filter(debit_note=i['id'], company=cmp).last()
    if p_history:
        i['action'] = p_history.action
        i['name']= f"{p_history.user.first_name} {p_history.user.last_name}"
        i['party'] = p_history.debit_note .party
    else:
        # Handle the case when no history is found
        i['action'] = ""
        i['name'] = ""
        i['party_name'] = ""


  return render(request, 'firstdebitnote.html',{'itm':itm,'dbill':dbill,'dbills':dbills,'usr':request.user})


def debit_note_redirect(request):
    # Check if any debit notes exist for the logged-in user
    debit_notes_exist = DebitNote.objects.filter(user=request.user).exists()

    if debit_notes_exist:
        # Debit notes exist, redirect to the page displaying the debit notes
        return redirect('debitnote2')  # Adjust the URL name as per your project

    else:
        # No debit notes exist, redirect to the page for creating the first debit note
        return redirect('firstdebitnote')

@csrf_exempt
def createdebitnote(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company
    
    parties = Party.objects.filter(company=cmp)
    items = Item.objects.filter(company=cmp)
    units = Unit.objects.filter(company=cmp)
    billno = PurchaseBill.objects.filter(company=cmp)

    try:
        max_reference_no = DebitNote.objects.filter(company=cmp).latest('returnno').returnno
        reference_no = max_reference_no + 1 if max_reference_no else 1
    except DebitNote.DoesNotExist:
        reference_no = 1

    context = {
        'usr': request.user,
        'parties': parties,
        'items': items,
        'units': units,
        'cmp': cmp,
        'reference_number': reference_no,
        'billno': billno
    }
    
    return render(request, 'createdebitnote.html', context)

@csrf_exempt
def purchasebilldata(request):
    if request.method == 'POST':
        if request.user.is_company:
            cmp = request.user.company
            print('cmp:',cmp)
        else:
            cmp = request.user.employee.company
            print('cmp1:',cmp)

        try:
            party_id = request.POST.get('id').split(" ")[0]
            print('party_id:',party_id)
            party = get_object_or_404(Party, company=cmp, id=party_id)
            print('party:',party)

            balance = party.openingbalance if party.openingbalance else None
            print('balance:',balance)
            phone = party.contact if party.contact else None
            print('phone:',phone)
            address = party.address if party.address else None
            print('address:',address)
            payment1 = party.payment if party.payment else None
            print('payment1:',payment1)

            # Initialize lists to store multiple bill numbers and dates
            billno = []
            print('billno:',billno)
            billdate = []
            print('billdate:',billdate)

            try:
                # Retrieve all instances for the party
                bill_instances = PurchaseBill.objects.filter(party=party, company=cmp)
                print('bill_instances:',bill_instances)

                # Get invoice numbers already saved in the CreditNote table
                debitnote_bill_numbers = DebitNote.objects.filter(bill__in=bill_instances).values_list('bill__billno', flat=True)
                print('cmp1:',cmp)

                # Loop through each instance and collect bill numbers and dates, excluding those in CreditNote
                for bill_instance in bill_instances:
                    if bill_instance.billno not in debitnote_bill_numbers:
                        billno.append(bill_instance.billno)
                        billdate.append(bill_instance.billdate)
                        print('bill_instance.billno:',bill_instance.billno)
                        print('cbill_instance.billdate:',bill_instance.billdate)

            except Invoice.DoesNotExist:
                pass

            # Return a JSON response with the list of bill numbers and dates
            if not billno and not billdate:
                return JsonResponse({
                    'billno': ['no bill'],
                    'billdate': ['nodate'],
                    'address': ['no address'],
                    'payment1': ['no payment'],
                    'phone': ['nophone']
                })

            return JsonResponse({
                'billno': billno,
                'billdate': billdate,
                'address': address,
                'bal': balance,
                'phone': phone,
                'payment1': payment1,
                'id': party.id
            })

        except KeyError:
            return JsonResponse({'error': 'The key "id" is missing in the POST request.'})

        except Party.DoesNotExist:
            return JsonResponse({'error': 'Party not found.'})

@csrf_exempt
def create_party(request):
    if request.user.is_authenticated:
        if request.method == "POST":
            if request.user.is_company:
                company_id = request.user.company.id
            else:
                company_id = request.user.employee.company.id
            user_id = request.user.id  # or request.user.pk

            party_name = request.POST.get('partyname')
            trn_no = request.POST.get('trn_no')
            contact = request.POST.get('contact')
            trn_type = request.POST.get('trn_type')
            state = request.POST.get('state')
            address = request.POST.get('address')
            email = request.POST.get('email')
            openingbalance = request.POST.get('balance')
            payment = request.POST.get('paymentType')
            current_date = request.POST.get('currentdate')
        
            additionalfield1 = request.POST.get('additionalfield1')
            additionalfield2 = request.POST.get('additionalfield2')
            additionalfield3 = request.POST.get('additionalfield3')

            new_party = Party(
                company_id=company_id,
                user_id=user_id,
                party_name=party_name,
                trn_no=trn_no,
                contact=contact,
                trn_type=trn_type,
                state=state,
                address=address,
                email=email,
                openingbalance=openingbalance,
                payment=payment,
                current_date=current_date,
                additionalfield1=additionalfield1,
                additionalfield2=additionalfield2,
                additionalfield3=additionalfield3,
            )
            new_party.save()

            parties = Party.objects.filter(company_id=company_id, user_id=user_id)
            data = {'name': new_party.party_name, 'id': new_party.id}
            return JsonResponse({'status': 'success', 'parties': data})
        else:
            return render(request, 'createdebitnote.html')
    else:
        return JsonResponse({'status': 'error', 'message': 'User is not authenticated'})

def extract_percentage(vat_string):
    # Split the string by space
    parts = vat_string.split()

    # Check if there are at least two parts (e.g., "VAT 5%")
    if len(parts) >= 2:
        # Get the second part and remove the percentage sign
        percentage = parts[1].replace('%', '')

        # Return the numeric value
        return int(percentage)

    # Return None if the string doesn't match the expected format
    return None


def save_debit_note(request):
  if request.method == 'POST':
    if request.user.is_company:
      cmp = request.user.company
      print('a:',cmp)
    else:
      cmp = request.user.employee.company
      print('a1:',cmp)
      
    usr = CustomUser.objects.get(username=request.user)
    print("usr:",usr)
    return_date=request.POST.get('returndate')
    print("returndate:",return_date)
    reference_no=request.POST.get('refnum')
    print("reference_no", reference_no)
    subtotal=request.POST.get('sub_total')
    print("sub: ",subtotal)
    vat=request.POST.get('disvatper')
    print("VAT: ",vat)
    adjustment=request.POST.get('adjustment')
    print("adj:",adjustment)
    grandtotal=request.POST.get('grandtotal')
    print("grandtotal: ",grandtotal)
    # party_status = request.POST.get('partystatus')
    # print("Partystatus: ",party_status)
    # # 
    debitnote_curr = DebitNote(
      user=usr, 
      company=cmp,
      returnno=reference_no,
      #artystatus=party_status,
      created_at=return_date,
      subtotal=subtotal,
      VAT=vat,
      adjustment=adjustment,
      grandtotal=grandtotal
      )
    debitnote_curr.save()
    print("debitnote_curr:",debitnote_curr)
    #if party_status=='partyon':
    party_details = request.POST.get('party_details')
    print("party_details:",party_details)
    party_id = party_details.split()[0]
    print("party_id:",party_id)
    party = Party.objects.get(pk=party_id,company=cmp)
      
    print(party.party_name)
    debitnote_curr.party=party
    debitnote_curr.save()
    print("debitnote_curr.party:",debitnote_curr.party)
      
    purchbill=request.POST.get('billno')
    print("purchbill:",purchbill)
    idsales=purchbill.split()[0]
    print("idsales:",idsales)
    pbill=PurchaseBill.objects.get(billno=idsales,company=cmp)
    print('pbill',pbill.billno)
    debitnote_curr.bill=pbill
    debitnote_curr.save()
     
    item_name =request.POST.getlist('item_name[]')
    print("item name: ",item_name)
    quantity =request.POST.getlist('qty[]')
    print("item qty: ",quantity)
    price =request.POST.getlist('price[]')
    print("item price",price)
    tax =request.POST.getlist('tax[]')
    print("item tax",tax)
    discount = request.POST.getlist('discount[]')
    print("item discount",discount)
    hsn = request.POST.getlist('hsn[]')
    print("item hsn",hsn)
    total = request.POST.getlist('total[]')
    print("item total",total)
    
    if len(quantity) == len(price) == len(tax) == len(discount) == len(hsn) == len(total) and item_name and quantity and price and tax and discount and hsn and total:
      mapped=zip(item_name,quantity,price,tax,discount,hsn,total)
      mapped=list(mapped)
      for ele in mapped:
        print("Element:", ele)
        hsn = ele[5]
        quantity = ele[1]
        tax = ele[3]
        price = ele[2]
        discount = ele[4]
        total = ele[6]
        print("HSN:", hsn)
        print("Quantity:", quantity)
        print("Tax:", tax)
        print("Price:", price)
        print("Discount:", discount)
        print("Total:", total)

        item_name_parts = ele[0].split()
        print("item_name_parts: ",item_name_parts)
        item_id = item_name_parts[0]
        print("item_id: ",item_id)
        items = Item.objects.get(id=item_id,company=cmp)
        print("items are: ",items)
        it=Item.objects.get(company=cmp, id = item_id).itm_name
        print("item_name:", it)
        # creditnote = CreditNote.objects.get(reference_no=creditnote_curr.reference_no,company=cmp)
        # print('1:',creditnote)
        debitnoteitem=DebitNoteItem(
                  user=usr,
                  debitnote=debitnote_curr,
                  company=cmp,
                  items=items,
                  item=it,
                  hsn=ele[5],
                  qty=ele[1],
                  VAT=ele[3],
                  price=ele[2],
                  discount=ele[4],
                  total=ele[6])
        debitnoteitem.save()
        print("debitnoteitem:", debitnoteitem)
   
      itt = ItemTransactions.objects.filter(item=items.id).last().trans_current_qty
      # Add invoice details in items transactions
      transaction = ItemTransactions.objects.create(
          company=cmp,
          item=items,
          trans_type='Debitnote',
          trans_date=debitnote_curr.created_at,
          trans_qty=ele[1],
          trans_current_qty = itt - int(ele[1]),
          trans_price=ele[2],
          #trans_invoice=invoice_curr.invoice_no
      )
      items.itm_stock_in_hand-= int(ele[1])
      items.save()
      #transaction.trans_current_qty -= int(ele[1]) 
      #transaction.save()
      print("Transaction saved:", transaction)
    
    p_id  = Party.objects.get(id=request.POST.get('partyname'))
    print('transaction0')
    p_id.openingbalance=float(p_id.openingbalance)-float(debitnote_curr.grandtotal)
    p_id.save()
    transactionparty=Transactions_party.objects.create(
      company=cmp,
      user=usr,
      party=p_id,
      trans_type='Debitnote',
      trans_number=debitnote_curr.returnno,
      trans_date=debitnote_curr.created_at,
      total=debitnote_curr.grandtotal,
      balance=p_id.openingbalance,
      
    )
    print(transactionparty.party)   
    print("Transaction party1:", transactionparty)
    transactionparty.save()


    debitnote = DebitNote.objects.get(returnno=debitnote_curr.returnno,company=cmp)
    print('1:',debitnote)
    cr=DebitNoteHistory(user=usr,company=cmp,debit_note=debitnote,action='Created')
    cr.save()
    if 'save_new' in request.POST:
      return redirect('createdebitnote')
    else:
      return redirect('firstdebitnote')

def debitnote2(request):
    # Assuming company id and user id are required for context
    company_id = request.session.get('company')
    user_id = request.session.get('user')

    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company

    print("Company Name:", cmp.company_name)  # Print the company name to the console

    parties = Party.objects.filter(company=cmp)
    items = Item.objects.filter(company=cmp)
    debits = DebitNote.objects.filter(user=request.user).prefetch_related('debitnotehistory_set')

    # Pass necessary data to the template context
    context = {
        'parties': parties,
        'items': items,
        'debits': debits,
        'company_id': company_id,
        'user_id': user_id,
        'company_name': cmp.company_name,  # Pass the company name to the template
        'usr': request.user  # Pass the user object to the template
    }
    
    return render(request, 'debitnote2.html', context)


def delete_debit_note(request, debitnote_id):
    if request.method == 'POST':
        try:
            # Retrieve the DebitNote object to be deleted
            debit_note = get_object_or_404(DebitNote, id=debitnote_id)
            
            # Delete related DebitNoteItem objects
            debit_note.debitnoteitem_set.all().delete()

            # Delete the DebitNote object
            debit_note.delete()

            # Return a success response
            return JsonResponse({'status': 'success'})

        except Exception as e:
            # If any error occurs, return an error response
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    else:
        # If the request method is not POST, return a method not allowed response
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
  
def search_debitnotes(request):
    if request.method == 'GET':
        fromDate_str = request.GET.get('fromDate')
        toDate_str = request.GET.get('toDate')

       
        fromDate = timezone.make_aware(datetime.strptime(fromDate_str, '%Y-%m-%d'))
        toDate = timezone.make_aware(datetime.strptime(toDate_str, '%Y-%m-%d'))

 
        debitnotes = DebitNote.objects.filter(created_at__range=[fromDate, toDate])

        
        debitnotes_list = list(debitnotes.values())

        
        return JsonResponse(debitnotes_list, safe=False)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    

def delete_debit_note_item(request,id):
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
  dbill = DebitNote.objects.get(id=id)
  DebitNoteItem.objects.filter(debitnote=dbill,company=cmp).delete()
  dbill.delete()
  
  return redirect('firstdebitnote')


def edit_debit_note(request, debit_id):

    # Fetch the company based on the user's role
    if request.user.is_company:
        cmp = request.user.company
        print('cmp1',cmp)
    else:
        cmp = request.user.employee.company
        print('cmp2',cmp)

    debitnote = get_object_or_404(DebitNote, id=debit_id)
    print('debitnote',debitnote)

   

    if request.method == 'POST':
        party_id = request.POST.get('party')
        print('party_id',party_id)
        
        return_no = request.POST.get('return_no')
        print('return_no',return_no)
        
        current_date = request.POST.get('current_date')
        print('current_date',current_date)
        subtotal = request.POST.get('subtotal')
        print('subtotal',subtotal)
        tax_amount = request.POST.get('taxAmount')
        print('tax_amount',tax_amount)
        adjustment = request.POST.get('adjustment')
        print('adjustment',adjustment)
        grand_total = request.POST.get('grandTotal')
        print('grand_total',grand_total)

        selected_party = Party.objects.get(id=party_id)
        print('selected_party',selected_party)

        # Update debit note fields
        debitnote.party = selected_party
        print('debitnote.party',debitnote.party)
        debitnote.returnno = return_no
        print('debitnote.returnno',debitnote.returnno)
        debitnote.created_at = current_date
        print('debitnote.created_at',debitnote.created_at)
        debitnote.subtotal = subtotal
        print('debitnote.subtotal',debitnote.subtotal)
        debitnote.taxamount = tax_amount
        print('debitnote.taxamount',debitnote.taxamount)
        debitnote.adjustment = adjustment
        print('debitnote.adjustment',debitnote.adjustment)
        debitnote.grandtotal = grand_total
        print('debitnote.grandtotal',debitnote.grandtotal)
        debitnote.save()
        print('debitnote',debitnote)

        # Update debit note items
        # items = request.POST.getlist("selected_item[]")
        items = request.POST.getlist("item_name[]")
        print('items',items)
        # quantities = request.POST.getlist("item_quantity[]")
        quantities = request.POST.getlist("qty[]")
        print('quantities',quantities)
        discounts = request.POST.getlist("discount[]")
        print('discounts',discounts)
        # total_amounts = request.POST.getlist("item_total_amount[]")
        total_amounts = request.POST.getlist("total[]")
        print('total_amounts',total_amounts)

        # Clear existing debit note items
        debitnote.debitnoteitem_set.all().delete()
        
        for item_id, qty, discount, total_amount in zip(items, quantities, discounts, total_amounts):
            itm = Item.objects.get(id=item_id)
            DebitNoteItem.objects.create(
                user=request.user,
                company_id=cmp.id,  # Pass the ID of the company instead of the company object
                debitnote=debitnote,
                items=itm,
                qty=qty,
                discount=discount,
                total=total_amount,
            )

        # Log the update action
        DebitNoteHistory.objects.create(
            user=request.user,
            company_id=cmp.id,
            debit_note=debitnote,
            date=timezone.now(),  # Use current timestamp
            action='Updated'
        )

        return redirect('firstdebitnote')  # Redirect to debit note list page or any other desired page

    else:  # Handling GET request
        parties = Party.objects.filter(company_id=cmp)
        print('parties',parties)
        items = Item.objects.filter(company_id=cmp)
        print('items',items)
        unit = Unit.objects.filter(company=cmp)
        print('unit',unit)
        debits = DebitNote.objects.filter(user=request.user).prefetch_related('debitnotehistory_set')
        print('debits',debits)
        current_date = debitnote.created_at.strftime('%Y-%m-%d') if debitnote.created_at else timezone.now().strftime('%Y-%m-%d')
        print('current_date',current_date)
        bills = PurchaseBill.objects.filter(party=debitnote.party)
        print('bills',bills)
        user_id = request.session.get('user')
        print('user_id',user_id)

        # Get selected items for the debit note
        selected_items = debitnote.debitnoteitem_set.all()
        print('selected_items',selected_items)
        selected_item = []
        print('selected_item',selected_item)
        for item in selected_items:
            selected_item.append({
                'id': item.items.id,
                'name': item.items.itm_name,
                'itm_hsn': item.items.itm_hsn,
                'purchaseprice': item.items.itm_purchase_price,
                'vat': item.items.itm_vat,
                'qty': item.qty,
                'discount': item.discount,
                'total_amount': item.total
            })
        print(selected_items)
        associated_bill = PurchaseBill.objects.filter(party=debitnote.party).first()
        print('associated_bill',associated_bill)

        context = {
            'debitnote': debitnote,
            'parties': parties,
            'items': items,
            'debits': debits,
            'company_id': cmp,
            'unit':unit,
            'user_id': user_id,
            'usr': request.user,
            'selected_party': debitnote.party.id if debitnote.party else None,
            'return_no': debitnote.returnno,
            'current_date': current_date,
            'subtotal': debitnote.subtotal,
            'taxAmount': debitnote.taxamount,
            'adjustment': debitnote.adjustment,
            'grandTotal': debitnote.grandtotal,
            'selected_party_address': debitnote.party.address if debitnote.party else None,
            'selected_party_contact': debitnote.party.contact if debitnote.party else None,
            'selected_party_balance': debitnote.party.openingbalance if debitnote.party else None,
            'billno': associated_bill.billno if associated_bill else None,
            'billdate': associated_bill.billdate if associated_bill else None,
            'selected_item': selected_item,
            'bills': bills,
            # Include additional context variables here
            'partyname': debitnote.party.party_name if debitnote.party else None,
            'balance': debitnote.party.openingbalance if debitnote.party else None,
            'contactnumber': debitnote.party.contact if debitnote.party else None,
            'bills': PurchaseBill.objects.filter(party=debitnote.party) if debitnote.party else None,
        }

        return render(request, 'editdebitnote.html', context)
    
def get_debit_note_history(request, debitnote_id):
    # Query the database for history data associated with the given debitnote_id
    history_data = DebitNoteHistory.objects.filter(debit_note__id=debitnote_id).select_related('debit_note__user').values('debit_note__returnno', 'date', 'user__first_name', 'user__last_name', 'action')

    # Convert queryset to list of dictionaries
    history_data_list = list(history_data)
    print("History data:", list(history_data))

    # Return the history data as JSON response
    return JsonResponse(history_data_list, safe=False)


# def generate_pdf(request, debit_note_id):
#     debit_note = DebitNote.objects.get(id=debit_note_id)
#     # Fetch related data
#     party = debit_note.party
#     debit_note_items = debit_note.debitnoteitem_set.all()  # Assuming you have related_name set for ForeignKey
#     # Render PDF template with fetched data
#     context = {
#         'debit_note': debit_note,
#         'party': party,
#         'debit_note_items': debit_note_items,
#     }
#     pdf_content = render_to_string('debit_note_pdf_template.html', context)

#     # Generate PDF using a library like WeasyPrint or ReportLab (not implemented here)
#     # For demonstration purposes, let's assume pdf_content contains the PDF data

#     # Return PDF as HttpResponse
#     response = HttpResponse(pdf_content, content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="debit_note_report.pdf"'
#     return response

def get_debit_note_details(request, debit_id):
    try:
        # Retrieve company based on user role
        if request.user.is_company:
            company = request.user.company
        else:
            company = request.user.employee.company

        # Retrieve debit note and related items
        debit_note = DebitNote.objects.get(id=debit_id)
        debit_note_items = DebitNoteItem.objects.filter(debitnote=debit_note)

        # Retrieve parties and items related to the company
        parties = Party.objects.filter(company=company)
        items = Item.objects.filter(company=company)

        # Prepare context data
        context = {
            'debit_id': debit_id,
            'usr': request.user,
            'company': company,
            'party': debit_note.party,
            'debit_note': debit_note,
            'debit_note_items': debit_note_items,
            'parties': parties,
            'items': items,
        }

        # Render template with context data
        return render(request, 'showtemplates.html', context)

    except (DebitNote.DoesNotExist, ValueError):
        # Return a JSON response with an error message if the debit note does not exist or if there's a value error
        return JsonResponse({'error': 'Invalid DebitNote ID'}, status=404)

    
@csrf_exempt  
def share_debit_note_via_email(request,id):
    if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                if request.user.is_company:
                  cmp = request.user.company
                else:
                  cmp = request.user.employee.company  
                usr = CustomUser.objects.get(username=request.user)
                # print(emails_list)
                debit_note = DebitNote.objects.get(id=id,company=cmp)
                debit_note_items = DebitNoteItem.objects.filter(debitnote=debit_note)
                dis = 0
                for itm in debit_note_items:
                  dis += int(itm.discount)
                itm_len = len(debit_note_items)
                context={'debit_note':debit_note,'debit_note_items':debit_note_items,'itm_len':itm_len,'dis':dis}
                template_path = 'debit_note_pdf_template.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                filename = f'Debitnote - {debit_note.returnno}.pdf'
                subject = f"Debitnote - {debit_note.returnno}"
                email = EmailMessage(subject, f"Hi,\nPlease find the Invoice bill- Bill-{debit_note.returnno}. \n{email_message}\n\n--\nRegards,\n{debit_note.company.company_name}\n{debit_note.company.address}\n - {debit_note.company.city}\n{debit_note.company.contact}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Debitnote has been shared via email successfully..!')
                return redirect(get_debit_note_details,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(get_debit_note_details, id)


def get_bill_date(request):
    user_id = request.user.id  # Assuming user is authenticated and user_id is obtained from the request
    print('1:',user_id)
    billno = request.GET.get('billno', None)
    print('2:',billno)
    
    if not billno:
        return JsonResponse({'error': 'Bill number not provided'}, status=400)
    
    try:
        # Query the PurchaseBill with the specified billno and user_id
        bill = PurchaseBill.objects.get(billno=billno, staff_id=user_id)
        print('3:',bill)
        billdate = bill.billdate.strftime('%Y-%m-%d')
        print('4:',billdate)
        return JsonResponse({'billdate': billdate})
    except PurchaseBill.DoesNotExist:
        return JsonResponse({'error': 'Bill number not found or you do not have access to this bill'}, status=404)
    except PurchaseBill.MultipleObjectsReturned:
        return JsonResponse({'error': 'Multiple bills found for the same bill number'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)

def additional_party_details(request):
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        party_id = request.POST.get('id')
        try:
            party = Party.objects.get(id=party_id)
            # Assuming Party model has fields: contact, address, balance
            data = {
                'contact': party.contact,
                'address': party.address,
                'balance': party.openingbalance,
            }
            print(data)
            return JsonResponse(data)
        except Party.DoesNotExist:
            return JsonResponse({'error': 'Party not found'}, status=404)
    else:
        return JsonResponse({'error': 'Invalid request'}, status=400)
        
        
def item_create1(request):
    if request.method == 'POST':
        print('Request received')

        # Retrieve each variable from the request
        itm_type = request.POST.get('itm_type')
        print('Item Type:', itm_type)

        itm_name = request.POST.get('itemName')
        print('Item Name:', itm_name)
        if Item.objects.filter(itm_name=itm_name).exists():
          response_data = {'success': False, 'message': 'Item name already exists!'}
          return JsonResponse(response_data)


        itm_hsn = request.POST.get('itemHSN')
        print('HSN:', itm_hsn)

        if Item.objects.filter(itm_hsn=itm_hsn).exists():
            response_data = {'success': False, 'message': 'HSN already exists!'}
            return JsonResponse(response_data)

        itm_unit = request.POST.get('unit')
        print('Unit:', itm_unit)

        itm_taxable = request.POST.get('taxable_result')
        print('Taxable:', itm_taxable)

        itm_vat = extract_percentage(request.POST.get('vat1'))
        print('VAT:', itm_vat)

        itm_sale_price = request.POST.get('sal_price')
        print('Sales Price:', itm_sale_price)

        itm_purchase_price = request.POST.get('pur_price')
        print('Purchase Price:', itm_purchase_price)

        itm_stock_in_hand = request.POST.get('opn_stock_')
        print('Stock In Hand:', itm_stock_in_hand)

        itm_at_price = request.POST.get('at_price')
        print('AT Price:', itm_at_price)

        itm_date = request.POST.get('date')
        print('Date:', itm_date)

        # Create item object and save
        item = Item(
            user=request.user,
            company=request.user.company,
            itm_type=itm_type,
            itm_name=itm_name,
            itm_hsn=itm_hsn,
            itm_unit=itm_unit,
            itm_taxable=itm_taxable,
            itm_vat=itm_vat,
            itm_sale_price=itm_sale_price,
            itm_purchase_price=itm_purchase_price,
            itm_stock_in_hand=itm_stock_in_hand,
            itm_at_price=itm_at_price,
            itm_date=itm_date
        )
        item.save()

       
        response_data = {'success': True, 'message': 'Item created successfully!', 'item_id': item.id, 'itemName': item.itm_name}
        return JsonResponse(response_data)

    # Render a response if not a POST request
    return render(request, 'createdebitnote.html')


def create_unit1(request):
    if request.method == 'POST':
        unit_name = request.POST.get('unit_name', '')
        company_id = request.user.company.id
        print(f"Company ID: {company_id}")

        company = Company.objects.get(id=request.user.company.id)

        # Create a new Unit instance
        unit = Unit.objects.create(
            company=company,
            unit_name=unit_name
        )
        
        # Prepare the JSON response data
        response_data = {
            'success': True,
            'message': 'Unit created successfully!',
            'unit_name': unit.unit_name,
            'unit_id': unit.id
        }
        print("Response Data:", response_data)

        return JsonResponse(response_data)

    # Handle other HTTP methods if needed
    return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    
def all_invoice(request):
  if request.user.is_company:
          cmp = request.user.company
  else:
          cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
   
  itm=Invoice.objects.filter(company=cmp)
  inbill = Invoice.objects.filter(company=cmp).values()
  inbills = Invoice.objects.filter(company=cmp)

  for i in inbill:
    p_history = InvoiceHistory.objects.filter(invoice_history=i['id'], company=cmp).last()
    if p_history:
        i['action'] = p_history.action
        i['name']= f"{p_history.user.first_name} {p_history.user.last_name}"
        i['party'] = p_history.invoice_history.party
    else:
        # Handle the case when no history is found
        i['action'] = ""
        i['name'] = ""
        i['party_name'] = ""


  return render(request, 'all_invoicedetails.html',{'itm':itm,'inbill':inbill,'inbills':inbills,'usr':request.user})

def create_invoice(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company
    parties = Party.objects.filter(company=cmp)
    items = Item.objects.filter(company=cmp)
    unit = Unit.objects.filter(company=cmp)
    
    try:
      max_reference_no = Invoice.objects.filter(company=cmp).last().reference_no
    except AttributeError:
      max_reference_no = None

    if max_reference_no is not None:
        reference_no = max_reference_no + 1
    else:
        reference_no = 1
# Check if there are deleted credit notes in InvoiceReference
    if InvoiceReference.objects.filter(company=cmp).exists():
      last_reference_no = InvoiceReference.objects.filter(company=cmp).last().reference_no
      print("last_reference_no", last_reference_no)
    inv = InvoiceReference.objects.filter(company=cmp).last()
    context = {'usr':request.user, 'parties':parties, 'items':items,'unit':unit,'cmp':cmp,'reference_number': reference_no,'inv':inv}
    
    return render(request, 'create_invoice.html', context)

def savePartyinvoice(request):
    if request.user.is_authenticated:
      if request.method == "POST":
          if request.user.is_company:
            cmp = request.user.company
          else:
            cmp = request.user.employee.company  
          print(cmp)
          usr = CustomUser.objects.get(username=request.user)
          party_name = request.POST['party_name']
          gst_no = request.POST['gst_no']
          mob = request.POST['party_num']
          gsttype = request.POST['gsttype']
          #state = request.POST['state']
          email = request.POST['email']
          addr = request.POST['party_addr']
          opbal = request.POST['creditamt']
          payment = request.POST.get('paymentType','')
          date = request.POST['credit_date']
          add1 = request.POST['addField1']
          add2 = request.POST['addField2']
          add3 = request.POST['addField3']
          if Party.objects.filter(trn_no=gst_no, company=cmp).exists() and not(gsttype == 'Unregistered/Consumers'):
            error_message = 'TRN number already exists!!!'
            return JsonResponse({'error_message': error_message})
          if gst_no == '' and (gsttype == 'Registered Business - Regular' or gsttype == 'Registered Business - Composition'):
            error_message = 'TRN number required.'
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(email=email, company=cmp).exists():
            error_message = 'Email already exists!!!'
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(contact=mob,company=cmp).exists():
            error_message = 'Mobile number already exists!!!'
            return JsonResponse({'error_message': error_message})
          user = request.user  
          party = Party(
                  user=user,
                  company=cmp,
                  party_name=party_name,
                  trn_no=gst_no,
                  contact=mob,
                  trn_type=gsttype,
                  #state=state,
                  address=addr,
                  email=email,
                  openingbalance=opbal,
                  payment=payment,
                  current_date=date,
                  additionalfield1=add1,
                  additionalfield2=add2,
                  additionalfield3=add3
              )
          party.save()
          print('Party created succefully ')
          trans = Transactions_party(user=request.user,company=cmp, trans_type='Opening Balance', trans_number=gst_no,
                                           trans_date=date, total=opbal, balance=opbal, party=party)
          trans.save()
          tr_history = PartyTransactionHistory(party=party, Transactions_party=trans, action="CREATED")
          tr_history.save()
          return redirect('create_invoice')

def party_dropdown_invoice(request):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company  
  options={}
  option_objects=Party.objects.filter(company=cmp)
  for option in option_objects:
    options[option.id]=option.party_name
  return JsonResponse(options)
 
def check_journal_num_valid(request):
    journals =InvoiceReference.objects.filter(pattern__startswith=str(request.user.id))
    journal_recieved_number = request.POST.get('journal_no')
    print(f'================== journal_recieved_number = {journal_recieved_number}==================')
    if journals.exists():
        last = journals.last()
        last_id = last.inv_rec_number
        print(f'================== last_id = {last_id}==================')
        if journal_recieved_number == last_id:
            return True
        else:
            return False
    else:
       
        return True
        
        
def saveinvoice(request):
  if request.method == 'POST':
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company
    usr = CustomUser.objects.get(username=request.user)
    
    #return_date=request.POST.get('returndate')
    #print("date",return_date)
    
    reference_no=request.POST.get('refnum')
    print("reference_no", reference_no)
    
    subtotal=request.POST.get('subtotal')
    print("subtotal",subtotal)
    
    vat=request.POST.get('disvatper')
    print("vat",vat)
    
    adjustment=request.POST.get('adjustment')
    print("adjustment",adjustment)
    
    grandtotal=request.POST.get('grandTotal')
    print("grandtotal",grandtotal)
    
    party_status = request.POST.get('partystatus')
    print("Partystatus: ",party_status)
    
    invoice_no = request.POST.get('invnum')
    
    inv = InvoiceReference.objects.filter(company=cmp).last()
    #last = ''
    #if inv.exists():
    #  last = inv.last()
    is_valid = check_journal_num_valid(request)
    print("good")
    print(is_valid)
    if not is_valid:
        messages.error(request, 'Invalid invoice number. Please enter a valid and continuous numeric sequence.')

    if InvoiceReference.objects.filter(user=request.user.id).exists():
        inv = InvoiceReference.objects.filter(user=request.user.id)
        inv_id = inv.last()
        inv_id1 = inv.last()

        # Check if there is a second last journal record
        if inv.exclude(id=inv_id.id).last():
            inv_id_second_last = inv.exclude(id=inv_id.id).last()
            pattern = inv_id_second_last.pattern
        else:
            inv_id_second_last = inv.first()
            pattern = inv_id_second_last.pattern

        if invoice_no != inv_id.inv_rec_number and invoice_no != '':
            # Creating a new JournalRecievedIdModel instance
            inv_id = InvoiceReference(user=request.user)
            count_for_ref_no =InvoiceReference.objects.filter(user=request.user.id).count()
            inv_id.pattern = pattern
            inv_id.save()

            # Using count_for_ref_no + 1 as the reference number
            ref_num = int(count_for_ref_no) + 2
            inv_id.reference_no = f'{ref_num:02}'

            inv_id.inv_rec_number = inv_id1.inv_rec_number
            inv_id.save()
        else:
            # Creating a new JournalRecievedIdModel instance
            inv_id =InvoiceReference(company=cmp)
            count_for_ref_no =InvoiceReference.objects.filter(user=request.user.id).count()
            inv_id.pattern = pattern
            inv_id.save()

            # Using count_for_ref_no + 1 as the reference number
            ref_num = int(count_for_ref_no) + 2
            inv_id.reference_no = f'{ref_num:02}'

            # Incrementing the jn_rec_number
            inv_rec_num = ''.join(i for i in inv_id1.inv_rec_number if i.isdigit())
            inv_rec_num = int(inv_rec_num)+1
            print("#################################")
            print(f"-----------------{inv_id1}-----------------")
            inv_id.inv_rec_number = f'{pattern}{inv_rec_num:02}'
            print(inv_id.inv_rec_number)
            inv_id.save()
            
    else:
        # Creating a new JournalRecievedIdModel instance
        inv_id = InvoiceReference(user=request.user)
        inv_id.save()

        # Setting initial values for ref_number, pattern, and jn_rec_number
        inv_id.reference_no = f'{2:02}'

        pattern = ''.join(i for i in invoice_no if not i.isdigit())
        inv_id.pattern = pattern
        inv_id.inv_rec_number = f'{pattern}{2:02}'
        inv_id.save()
    inv_id.company=cmp  

    #inv_id.user = request.session['login_id']
    inv_id.save()
    invoice_curr = Invoice(
      user=usr,
      company=cmp,
      reference_no=reference_no,
      partystatus=party_status,
      #returndate=return_date,
      invoice_no=invoice_no, 
      subtotal=subtotal,
      vat=vat,
      adjustment=adjustment, 
      grandtotal=grandtotal)
    invoice_curr.save()
    print("invoice_curr: ",invoice_curr)
    
    if party_status=='partyon':
      party_details = request.POST.get('party_details')
      party_id = party_details.split()[0]
      party = Party.objects.get(pk=party_id,company=cmp)
      
      print(party.party_name)
      invoice_curr.party=party
      invoice_curr.save()
     
    
    item_name =request.POST.getlist('item_name[]')
    print("item name: ",item_name)
    quantity =request.POST.getlist('qty[]')
    print("item qty: ",quantity)
    price =request.POST.getlist('price[]')
    print("item price",price)
    tax =request.POST.getlist('tax[]')
    print("item tax",tax)
    discount = request.POST.getlist('discount[]')
    print("item discount",discount)
    hsn = request.POST.getlist('hsn[]')
    print("item hsn",hsn)
    total = request.POST.getlist('total[]')
    print("item total",total)
    
    if len(item_name) == len(quantity) == len(price) == len(tax) == len(discount) == len(hsn) == len(total) and item_name and quantity and price and tax and discount and hsn and total:
      mapped=zip(item_name,quantity,price,tax,discount,hsn,total)
      mapped=list(mapped)
      for ele in mapped:
        print("Element:", ele)
        hsn = ele[5]
        quantity = ele[1]
        tax = ele[3]
        price = ele[2]
        discount = ele[4]
        total = ele[6]
        print("HSN:", hsn)
        print("Quantity:", quantity)
        print("Tax:", tax)
        print("Price:", price)
        print("Discount:", discount)
        print("Total:", total)

        item_name_parts = ele[0].split()
        print("item_name_parts: ",item_name_parts)
        item_id = item_name_parts[0]
        print("item_id: ",item_id)
        items = Item.objects.get(id=item_id,company=cmp)
        print("items are: ",items)
        it=Item.objects.get(company=cmp, id = item_id).itm_name
        print("item_name:", it)
        
        invoiceitem=InvoiceItem(
                  user=usr,
                  invoice=invoice_curr,
                  company=cmp,
                  items=items,
                  item=it,
                  hsn=ele[5],
                  quantity=ele[1],
                  tax=ele[3],
                  price=ele[2],
                  discount=ele[4],
                  total=ele[6])
        invoiceitem.save()
        print("invoiceitem:", invoiceitem)
    
    itt = ItemTransactions.objects.filter(item=items.id).last().trans_current_qty
    # Add invoice details in items transactions
    transaction = ItemTransactions.objects.create(
        company=cmp,
        item=items,
        trans_type='Invoice',
        trans_date=invoice_curr.invoice_date,
        trans_qty=ele[1],
        trans_current_qty = itt - int(ele[1]),
        trans_price=ele[2],
        #trans_invoice=invoice_curr.invoice_no
    )
    items.itm_stock_in_hand-= int(ele[1])
    items.save()
    #transaction.trans_current_qty -= int(ele[1]) 
    #transaction.save()
    print("Transaction saved:", transaction)
    
    p_id  = Party.objects.get(id=request.POST.get('partyname'))
    print('transaction0')
    p_id.openingbalance=float(p_id.openingbalance)-float(invoice_curr.grandtotal)
    p_id.save()
    transactionparty=Transactions_party.objects.create(
      company=cmp,
      user=usr,
      party=p_id,
      trans_type='Invoice',
      trans_number=invoice_curr.invoice_no,
      trans_date=invoice_curr.invoice_date,
      total=invoice_curr.grandtotal,
      balance=p_id.openingbalance,
      
    )
    print(transactionparty.party)   
    print("Transaction party1:", transactionparty)
    transactionparty.save()
      
    cr=InvoiceHistory(user=usr,company=cmp,invoice_history=invoice_curr,action='Created')
    print("cr", cr)
    cr.save()
    

    
    if 'save_new' in request.POST:
      return redirect('create_invoice')
    else:
      return redirect('all_invoice')
    
def details_invoicebill(request,id):
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
  inbill = Invoice.objects.get(id=id,company=cmp)
  initm = InvoiceItem.objects.filter(invoice=inbill,company=cmp)
  dis = 0
  for itm in initm:
    dis += int(itm.discount)
  itm_len = len(initm)

  context={'inbill':inbill,'initm':initm,'itm_len':itm_len,'dis':dis,'usr':request.user}
  return render(request,'invoicebilldetails.html',context)

def shareinvoicepdftomail(request,id):
 if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                if request.user.is_company:
                  cmp = request.user.company
                else:
                  cmp = request.user.employee.company  
                usr = CustomUser.objects.get(username=request.user)
                # print(emails_list)
                inbill = Invoice.objects.get(id=id,company=cmp)
                initm = InvoiceItem.objects.filter(invoice=inbill)
                dis = 0
                for itm in initm:
                  dis += int(itm.discount)
                itm_len = len(initm)
                context={'inbill':inbill,'initm':initm,'itm_len':itm_len,'dis':dis}
                template_path = 'invoicepdf.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                filename = f'Invoice bill - {inbill.invoice_no}.pdf'
                subject = f"Invoice bill - {inbill.invoice_no}"
                email = EmailMessage(subject, f"Hi,\nPlease find the Invoice bill- Bill-{inbill.invoice_no}. \n{email_message}\n\n--\nRegards,\n{inbill.company.company_name}\n{inbill.company.address}\n - {inbill.company.city}\n{inbill.company.contact}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Invoice bill has been shared via email successfully..!')
                return redirect(details_invoicebill,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(details_invoicebill, id) 
          
          
def history_invoicebill(request,id):
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company 
  usr = CustomUser.objects.get(username=request.user) 
  inbill = Invoice.objects.get(id=id)
  hst= InvoiceHistory.objects.filter(invoice_history=inbill,company=cmp)

  context = {'hst':hst,'inbill':inbill,'usr':request.user}
  return render(request,'invoicehistory.html',context) 

def delete_invoice(request,id):
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
  inbill = Invoice.objects.get(id=id)
  InvoiceItem.objects.filter(invoice=inbill,company=cmp).delete()
  inbill.delete()
  
  return redirect('all_invoice')

def invoicebillhistory(request):
  pid = request.POST['id']
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company
  usr = CustomUser.objects.get(username=request.user) 
  inbill = Invoice.objects.get(invoice_no=pid,company=cmp,staff=usr)
  hst = InvoiceHistory.objects.filter(invoice_history=inbill,company=cmp).last()
  name = hst.user.first_name + ' ' + hst.user.last_name 
  action = hst.action
  return JsonResponse({'name':name,'action':action,'pid':pid,'usr':request.user})
        

def edit_invoice(request,pk):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company
  parties = Party.objects.filter(company=cmp)
  items = Item.objects.filter(company=cmp)
  unit = Unit.objects.filter(company=cmp)
  invoice_curr=Invoice.objects.get(id=pk,company=cmp)
  reference=invoice_curr.reference_no
  invoice_items=InvoiceItem.objects.filter(invoice=invoice_curr,company=cmp)
  for item in invoice_items:
    print(f"Item ID: {item.id}")
    print(f"Credit Note: {item.item}")
  context={'usr':request.user,
           'invoice_curr':invoice_items,
           'invoice':invoice_curr,
           'parties':parties,
           'reference':reference,
           'items':items,'unit':unit
          }
  return render(request,'invoice_edit.html',context)


def updateinvoice(request,pk):
  if request.method=="POST":
    if request.user.is_company:
      cmp = request.user.company
    else:
      cmp = request.user.employee.company

    usr = CustomUser.objects.get(username=request.user)
    print(f"usr: {usr}")
    invoice=Invoice.objects.get(id=pk,company=cmp)
    print(f"invoice: {invoice}")
    invoice.user=request.user
    print(f"invoice.user: {invoice.user}")
    invoice.company=cmp
    print(f"invoice.company: {invoice.company}")
    #invoice.returndate=request.POST['returndate']
    #print(f"invoice.returndate: {invoice.returndate}")
    invoice.reference_no=request.POST['refnum']
    print(f"invoice.reference_no: {invoice.reference_no}")
    invoice.invoice_no=request.POST['invoiceno']
    print(f"invoice.invoice_no: {invoice.invoice_no}")
    invoice.subtotal=request.POST['subtotal']
    print(f"invoice.subtotal: {invoice.subtotal}")
    invoice.vat=request.POST['disvatper']
    print(f"invoice.vat: {invoice.vat}")
    invoice.adjustment=request.POST['adjustment']
    print(f"invoice.adjustment: {invoice.adjustment}")
    invoice.grandtotal=request.POST['grandTotal']
    print(f"invoice.grandtotal: {invoice.grandtotal}")
    invoice.partystatus = request.POST.get('partystatus')
    print(f"invoice.partystatus: {invoice.partystatus}")
    invoice.save()

    if invoice.partystatus=='partyon':
      party_details = request.POST.get('party_details')
      party_id = party_details.split()[0]
      party = Party.objects.get(pk=party_id,company=cmp)
      party.contact=request.POST.get('party_phone')
      party.payment=request.POST.get('paymethod')
      party.save()
      print(party.party_name)
      invoice.party=party
     
      invoice.save()
    else:
      invoice.party = None
      invoice.save()
    InvoiceHistory.objects.create(user=usr,company=cmp,invoice_history=invoice,action='Updated')


    item_name = request.POST.getlist('item_name[]')
    print(f"item_name: {item_name}")
    quantity = request.POST.getlist('qty[]')
    print(f"quantity: {quantity}")
    price = request.POST.getlist('price[]')
    print(f"price: {price}")
    tax = request.POST.getlist('tax[]')
    print(f"tax: {tax}")
    discount = request.POST.getlist('discount[]')
    print(f"discount: {discount}")
    hsn = request.POST.getlist('hsn[]')
    print(f"hsn: {hsn}")
    total = request.POST.getlist('total[]')
    print(f"total: {total}")
    
    InvoiceItem.objects.filter(invoice=invoice).delete()
    if len(item_name) == len(quantity) == len(price) == len(tax) == len(discount) == len(hsn) == len(total) and item_name and quantity and price and tax and discount and hsn and total:
      mapped=zip(item_name,quantity,price,tax,discount,hsn,total)
      mapped=list(mapped)
      #invoice_items = InvoiceItem.objects.filter(invoice=invoice,company=cmp)

      item_ids=[]
      for ele in mapped:
        print("forloop.counter")
        item_name_parts = ele[0].split()
        
        item_id = item_name_parts[0]
        item_ids.append(item_id)
        items = Item.objects.get(company=cmp,pk=item_id)
        it=Item.objects.get(company=cmp, id = item_id).itm_name
        print("item_name:", it)
        #existing_invoice_item = invoice_items.filter(items=items).first()
        InvoiceItem.objects.create(
                user=usr,
                invoice=invoice,
                company=cmp,
                items=items,
                item=it,
                hsn=ele[5],
                quantity=ele[1],
                tax=ele[3],
                price=ele[2],
                discount=ele[4],
                total=ele[6]
            )
        
      existing_item_ids = [int(id.split('_')[-1]) for id in item_ids]
      
      
    return redirect('all_invoice')
      
      
def details_party(request,id):
  if request.user.is_company:
    party = Party.objects.filter(company = request.user.company)
  else:
    party = Party.objects.filter(company = request.user.employee.company)

  fparty = Party.objects.get(id=id)
  ftrans = Transactions_party.objects.filter(party = fparty)
  context = {'party':party, 'usr':request.user, 'fparty':fparty, 'ftrans':ftrans}
  return render(request,'party_details.html',context)
  
  
def item_details(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    
    item_id = request.POST.get('id').split(" ")[0]
    item = Item.objects.get(company=cmp, pk=item_id)
    
    hsn = item.itm_hsn
    price = item.itm_sale_price
    tax = item.itm_vat
    
    return JsonResponse({'hsn': hsn, 'price': price, 'tax': tax})
    
    
def check_hsn_number_existsdebit(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    hsn = request.GET.get('hsn')
    if Item.objects.filter(itm_hsn=hsn, company=cmp).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists': False})
    
    
def savePartyinvoice1(request,pk):
    if request.user.is_authenticated:
      if request.method == "POST":
          if request.user.is_company:
            cmp = request.user.company
          else:
            cmp = request.user.employee.company  
          print(cmp)
          usr = CustomUser.objects.get(username=request.user)
          party_name = request.POST['party_name']
          gst_no = request.POST['gst_no']
          mob = request.POST['party_num']
          gsttype = request.POST['gsttype']
          #state = request.POST['state']
          email = request.POST['email']
          addr = request.POST['party_addr']
          opbal = request.POST['creditamt']
          payment = request.POST.get('paymentType','')
          date = request.POST['credit_date']
          add1 = request.POST['addField1']
          add2 = request.POST['addField2']
          add3 = request.POST['addField3']
          if Party.objects.filter(trn_no=gst_no, company=cmp).exists() and not(gsttype == 'Unregistered/Consumers'):
            error_message = 'TRN number already exists!!!'
            #return redirect('updateinvoice',pk=pk)
            return JsonResponse({'error_message': error_message})
          if gst_no == '' and (gsttype == 'Registered Business - Regular' or gsttype == 'Registered Business - Composition'):
            error_message = 'TRN number required.'
            #return redirect('updateinvoice',pk=pk)
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(email=email, company=cmp).exists():
            error_message = 'Email already exists!!!'
            #return redirect('updateinvoice',pk=pk)
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(contact=mob,company=cmp).exists():
            error_message = 'Mobile number already exists!!!'
            #return redirect('updateinvoice',pk=pk)
            return JsonResponse({'error_message': error_message})
          user = request.user  
          party = Party(
                  user=user,
                  company=cmp,
                  party_name=party_name,
                  trn_no=gst_no,
                  contact=mob,
                  trn_type=gsttype,
                  #state=state,
                  address=addr,
                  email=email,
                  openingbalance=opbal,
                  payment=payment,
                  current_date=date,
                  additionalfield1=add1,
                  additionalfield2=add2,
                  additionalfield3=add3
              )
          party.save()
          print('Party created succefully ')
          trans = Transactions_party(user=request.user,company=cmp, trans_type='Opening Balance', trans_number=gst_no,
                                           trans_date=date, total=opbal, balance=opbal, party=party)
          trans.save()
          tr_history = PartyTransactionHistory(party=party, Transactions_party=trans, action="CREATED")
          tr_history.save()
          return redirect('updateinvoice',pk=pk)
          
          
def Purchasereport(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company
        
    usr = CustomUser.objects.get(username=request.user)
      
    purchases = PurchaseBill.objects.filter(company=cmp)
    debit = DebitNote.objects.filter(company=cmp)
    parties = Party.objects.filter(company=cmp)
   
    total_purchase_amount = purchases.aggregate(total_purchase=Sum('grandtotal', output_field=DecimalField()))['total_purchase']
    total_debit_amount = debit.aggregate(total_debit=Sum('grandtotal', output_field=DecimalField()))['total_debit']

    if total_purchase_amount is not None:
        total_purchase_amount = round(total_purchase_amount, 2)

    if total_purchase_amount is not None and total_debit_amount is not None:
        total_amount_after_debit = total_purchase_amount - total_debit_amount
    else:
        total_amount_after_debit = None

    return render(request, 'Purchasereport.html', {
        'purchases': purchases,
        'debit': debit,
        'cmp': cmp,
        'parties': parties,
        'total_purchase_amount': total_purchase_amount,
        'total_debit_amount': total_debit_amount,
        'total_amount_after_debit': total_amount_after_debit,
        'usr': request.user
    })


def sharePurchaseReportsToEmail(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company
       
    if request.method == 'POST':
        emails_string = request.POST['email_ids']

        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(',')]
        email_message = request.POST['email_message']

        invoices = PurchaseBill.objects.filter(company=cmp)
        debit_notes = DebitNote.objects.filter(company=cmp)

        excelfile = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Purchases Reports'

        # Write headers
        headers = ['#', 'Date', 'Bill Number', 'Party Name', 'Type', 'Amount']
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        # Write Purchases invoices data
        for idx, invoice in enumerate(invoices, start=2):
            worksheet.cell(row=idx, column=1, value=idx - 1)  # Index
            worksheet.cell(row=idx, column=2, value=invoice.billdate.strftime('%Y-%m-%d'))  # Date
            worksheet.cell(row=idx, column=3, value=invoice.billno)  # Invoice Number
            worksheet.cell(row=idx, column=4, value=invoice.party.party_name)  # Party Name
            worksheet.cell(row=idx, column=5, value='Bill')  # Transaction Type
            worksheet.cell(row=idx, column=6, value=invoice.grandtotal)  # Amount
                    
        # Write credit notes data
        for idx, debit_note in enumerate(debit_notes, start=len(invoices) + 2):
            worksheet.cell(row=idx, column=1, value=idx - 1)  # Index
            worksheet.cell(row=idx, column=2, value=debit_note.created_at.strftime('%Y-%m-%d'))  # Date
            worksheet.cell(row=idx, column=3, value=debit_note.returnno)  # Return Number
            worksheet.cell(row=idx, column=4, value=debit_note.party.party_name)  # Party Name
            worksheet.cell(row=idx, column=5, value='Debit Note')  # Transaction Type
            worksheet.cell(row=idx, column=6, value=debit_note.grandtotal)  # Amount               

        # Save workbook to BytesIO object
        workbook.save(excelfile)
        mail_subject = f'Purchase Reports - {date.today()}'
        message = f"Hi,\nPlease find the ALES REPORTS file attached. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}"
     
        message = EmailMessage(mail_subject, message, settings.EMAIL_HOST_USER, emails_list)
        message.attach(f'Purchase Reports-{date.today()}.xlsx', excelfile.getvalue(), 'application/vnd.ms-excel')
        message.send(fail_silently=False)

        # messages.success(request, 'Purchase Report has been shared via email successfully..!')
        return redirect('Purchasereport')  
    # messages.error(request, 'An error occurred while sharing the purchase report via email.')
    return redirect('Purchasereport')


def Purchasereport_graph(request):
    if request.user:
        cmp = Company.objects.get(user=request.user.id)
        usr = CustomUser.objects.get(username=request.user)
        current_year = date.today().year  

        monthly_purchase_data = defaultdict(float)  
        for month in range(1, 13):
            monthly_purchase_data[month] = float(  
                PurchaseBill.objects
                .filter(billdate__month=month, billdate__year=current_year, company=cmp)
                .aggregate(total_purchase=Sum('grandtotal'))['total_purchase'] or 0             
            )
            
        monthly_debit_data = defaultdict(float)  
        for month in range(1, 13):
            monthly_debit_data[month] = float(  
                DebitNote.objects
                .filter(created_at__month=month, created_at__year=current_year, company=cmp)
                .aggregate(total_debit=Sum('grandtotal'))['total_debit'] or 0
            )

        yearly_purchase_data = defaultdict(float)  
        for year in range(2014, current_year + 1):
            yearly_purchase_data[year] = float(  
                PurchaseBill.objects
                .filter(billdate__year=year, company=cmp)
                .aggregate(total_purchase=Sum('grandtotal'))['total_purchase'] or 0
            )

        yearly_debit_data = defaultdict(float) 
        for year in range(2014, current_year + 1):
            yearly_debit_data[year] = float(  
                DebitNote.objects
                .filter(created_at__year=year, company=cmp)
                .aggregate(total_debit=Sum('grandtotal'))['total_debit'] or 0
            )

        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        monthly_labels = [f"{month_names[month - 1]} {current_year}" for month in range(1, 13)]
        monthly_purchase = [monthly_purchase_data[month] for month in range(1, 13)]
        monthly_debit = [monthly_debit_data[month] for month in range(1, 13)]

        yearly_labels = [str(year) for year in range(2014, current_year + 1)]
        yearly_purchase = [yearly_purchase_data[year] for year in range(2014, current_year + 1)]
        yearly_debit = [yearly_debit_data[year] for year in range(2014, current_year + 1)]

        chart_data = {
            'monthly_labels': monthly_labels,
            'monthly_purchase': monthly_purchase,
            'monthly_debit': monthly_debit,
            'yearly_labels': yearly_labels,
            'yearly_purchase': yearly_purchase,
            'yearly_debit': yearly_debit
        }
        return render(request, 'purchase_graph.html', {'chart_data': chart_data, 'cmp': cmp, 'usr': request.user})
        
        
def shareTransactionitemToEmail(request,id):
 if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                if request.user.is_company:
                  cmp = request.user.company
                else:
                  cmp = request.user.employee.company  
                usr = CustomUser.objects.get(username=request.user)
                # print(emails_list)
                inbill = Item.objects.get(id=id,company=cmp)
                initm = ItemTransactions.objects.filter(item=inbill)
                #dis = 0
                #for itm in initm:
                  #dis += int(itm.discount)
                itm_len = len(initm)
                context={'inbill':inbill,'initm':initm,'itm_len':itm_len}
                template_path = 'itempdf.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                filename = f'Item bill - {inbill.itm_name}.pdf'
                subject = f"Item bill - {inbill.itm_name}"
                email = EmailMessage(subject, f"Hi,\nPlease find the Item bill- Bill-{inbill.itm_name}. \n{email_message}\n\n--\nRegards,\n{inbill.company.company_name}\n{inbill.company.address}\n - {inbill.company.city}\n{inbill.company.contact}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Item bill has been shared via email successfully..!')
                return redirect(item_list,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(item_list, id)
            
def details_item(request,id):
  if request.user.is_company:
    itm= Item.objects.filter(company = request.user.company)
  else:
    itm = Item.objects.filter(company = request.user.employee.company)

  fitem = Item.objects.get(id=id)
  fitems = ItemTransactions.objects.filter(item= fitem)
  context = {'itm':itm, 'usr':request.user, 'fitem':fitem, 'fitems':fitems}
  return render(request,'item_details.html',context)
  
  
@login_required(login_url='login')
def stock_report(request):    
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company
        
    usr = CustomUser.objects.get(username=request.user)
    print("usr:",usr)
    
    stockList = []
    items = Item.objects.filter(company= cmp)
    print("items:",items)
    
    for item in items:
        stockIn = 0
        stockOut = 0
        for i in ItemTransactions.objects.filter(company = cmp, item = item).filter(trans_type= 'Purchase'):
          print("iin",i)
          stockIn += i.trans_qty
          #stockIn = int(stockIn)+ int(i.trans_qty)
          print("stockin:",stockIn)

        for i in ItemTransactions.objects.filter(company = cmp, item = item).filter(trans_type = 'Invoice'):
          stockOut += i.trans_qty
          #stockOut = int(stockOut)+ int(i.trans_qty)
          print("stockout:",stockOut)

        dict = {
            'name':item.itm_name,
            'stockIn':stockIn,
            'stockOut':stockOut,
            'balance':item.itm_stock_in_hand
        }
        print("dict:",dict)
        
        stockList.append(dict)

    context = {
        'cmp':cmp,
        'items':items,
        'stock':stockList,
        'count':items.count(),
    }
    return render(request, 'stock_report.html',context)

def itemStockReport(request,id):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company
        
    usr = CustomUser.objects.get(username=request.user)
    print("usr:",usr)
        
    if id == 0:
        return redirect(stock_report)
    stockList = []
    items = Item.objects.get(company= cmp, id = id)
    print("item:",items)
    
    stockIn = 0
    stockOut = 0
    for i in ItemTransactions.objects.filter(company= cmp, item = items).filter(trans_type = 'Purchase'):
      stockIn += i.trans_qty
        #stockIn = int(stockIn)+ int(i.trans_qty)
      print("stockin:",stockIn)

    for i in ItemTransactions.objects.filter(company = cmp, item = items).filter(trans_type = 'Invoice'):
      
      stockOut += i.trans_qty
        #stockOut = int(stockOut)+ int(i.trans_qty)
      print("stockout:",stockOut)

    dict = {
        'name':items.itm_name,
        'stockIn':stockIn,
        'stockOut':stockOut,
        'balance':items.itm_stock_in_hand
    }
    
    stockList.append(dict)

    context = {
        'cmp':cmp,
        'items':Item.objects.filter(company= cmp),
        'stock':stockList,
        'balance':items.itm_stock_in_hand,
        'id':id,
    }
    return render(request, 'stock_report.html',context)

def sharestockitemToEmail(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company
       
    if request.method == 'POST':
        emails_string = request.POST['email_ids']

        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(',')]
        email_message = request.POST['email_message']

        items = Item.objects.filter(company=cmp)
        #Itemtransaction = ItemTransactions.objects.filter(company=cmp, item=items)

        excelfile = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Purchases Reports'

        # Write headers
        headers = ['#', 'ITEM', 'STOCK IN', 'STOCK OUT', 'BALANCE']
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        # Write Purchases invoices data
        for idx, item in enumerate(items, start=2):
            worksheet.cell(row=idx, column=1, value=idx - 1)  # Index
            worksheet.cell(row=idx, column=2, value=item.itm_name)  # item name
            worksheet.cell(row=idx, column=3, value=item.stockIn)  # stockin
            worksheet.cell(row=idx, column=4, value=item.stockOut)  # stockout
            worksheet.cell(row=idx, column=5, value=item.itm_stock_in_hand)  # balance
                          

        # Save workbook to BytesIO object
        workbook.save(excelfile)
        mail_subject = f'StockReports - {date.today()}'
        message = f"Hi,\nPlease find the ALES REPORTS file attached. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}"
     
        message = EmailMessage(mail_subject, message, settings.EMAIL_HOST_USER, emails_list)
        message.attach(f'Stock Reports-{date.today()}.xlsx', excelfile.getvalue(), 'application/vnd.ms-excel')
        message.send(fail_silently=False)

        #messages.success(request, 'Purchase Report has been shared via email successfully..!')
        return redirect('stock_report')  
    #messages.error(request, 'An error occurred while sharing the purchase report via email.')
    return redirect('stock_report')
    
    
def report(request):
  if request.user.is_company:
    cmp = request.user.company
    party = Party.objects.filter(company = request.user.company)

  else:
    cmp = request.user.employee.company

    party = Party.objects.filter(company = request.user.employee.company)
  
  if party:
    fparty = party[0]
    ftrans = Transactions_party.objects.filter(party = fparty)
    itm=Invoice.objects.filter(company=cmp)
    inbill = Invoice.objects.filter(company=cmp)
    credit=CreditNote.objects.all().filter(company=cmp)
    subtotal_a = Invoice.objects.filter(company=cmp).aggregate(total=Sum('grandtotal'))['total'] or 0
    subtotal_b = credit.aggregate(total=Sum('grandtotal'))['total'] or 0
    result = subtotal_a - subtotal_b
    

   
  
    context = {'party':party, 'usr':request.user, 'fparty':fparty, 'ftrans':ftrans,'itm':itm,'inbill':inbill,'cmp':cmp,'credit':credit,'result':result}
  else:
        context = {'party':party, 'usr':request.user}

  return render(request,'report.html',context)
 

def sendmail_report(request, id):
    if request.user.is_company:
        party = Party.objects.filter(company=request.user.company)
    else:
        party = Party.objects.filter(company=request.user.employee.company)
        
    if request.method == "POST":
        try:
         
            fparty = Party.objects.get(id=id)
            ftrans = Transactions_party.objects.filter(party=fparty)
            cmp = fparty.company
            itm=Invoice.objects.filter(company=cmp)
            credit=CreditNote.objects.all().filter(company=cmp)
            context = {'party': party, 'usr': request.user, 'fparty': fparty, 'ftrans': ftrans,'cmp':cmp,'itm':itm,'credit':credit}
            
            email_message = request.POST.get('email_message')
            my_subject = "SALES REPORT"
            emails_string = request.POST.get('email_ids')
            emails_list = [email.strip() for email in emails_string.split(',')]
            
            html_message = render_to_string('report_pdf.html', context)
            plain_message = strip_tags(html_message)
            
            pdf_content = BytesIO()
            pisa.CreatePDF(html_message.encode("UTF-8"), pdf_content)
            pdf_content.seek(0)

            filename = f'report.pdf'
            email=EmailMultiAlternatives(my_subject,f"Hi,\nPlease find the attached sales Report - \n{email_message}\n--\nRegards,\n{cmp.company_name},\n{cmp.address},{cmp.city},{cmp.country},\n{cmp.contact}\n",from_email='altostechnologies6@gmail.com',
                                         to=emails_list, )
       
            email.attach(filename, pdf_content.read(), 'application/pdf')
            email.send()

            return HttpResponse('<script>alert("Report has been shared successfully!");window.location="/report"</script>')
        except Party.DoesNotExist:
            return HttpResponse('<script>alert("Party not found!");window.location="/report"</script>')
        except Exception as e:
            # Handle the exception, log the error, or provide an error message
            return HttpResponse(f'<script>alert("Failed to send email: {str(e)}");window.location="/report"</script>')

    return HttpResponse('<script>alert("Invalid Request!");window.location="/report"</script>')
  
def sales_graph(request):

    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company
    

    invoices_monthly = Invoice.objects.filter(company=cmp).annotate(month=TruncMonth('invoice_date')).values('month').annotate(total=Sum('grandtotal'))
    credit_notes_monthly = CreditNote.objects.filter(company=cmp).annotate(month=TruncMonth('creditnote_date')).values('month').annotate(total=Sum('grandtotal'))


    monthly_difference = {month_name[i]: 0 for i in range(1, 13)}

    
    for month_data in invoices_monthly:
        month = month_data['month'].strftime('%B')
        monthly_difference[month] += float(month_data['total']) if month_data['total'] else 0

    for month_data in credit_notes_monthly:
        month = month_data['month'].strftime('%B')
        monthly_difference[month] -= float(month_data['total']) if month_data['total'] else 0


    monthly_labels = list(monthly_difference.keys())
    monthly_data = list(monthly_difference.values())


    invoices_yearly = Invoice.objects.filter(company=cmp).annotate(year=TruncYear('invoice_date')).values('year').annotate(total=Sum('grandtotal'))
    credit_notes_yearly = CreditNote.objects.filter(company=cmp).annotate(year=TruncYear('creditnote_date')).values('year').annotate(total=Sum('grandtotal'))

    
    min_year = 2021
    max_year = max(
        max((data['year'].year for data in invoices_yearly), default=min_year),
        max((data['year'].year for data in credit_notes_yearly), default=min_year)
    )
    years = list(range(min_year, max_year + 1))

   
    yearly_difference = {year: 0 for year in years}


    for year_data in invoices_yearly:
        year = year_data['year'].year
        yearly_difference[year] += float(year_data['total']) if year_data['total'] else 0

    for year_data in credit_notes_yearly:
        year = year_data['year'].year
        yearly_difference[year] -= float(year_data['total']) if year_data['total'] else 0


    yearly_labels = list(yearly_difference.keys())
    yearly_data = list(yearly_difference.values())


    return render(request, 'sales_graph.html', {
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'yearly_labels': json.dumps(yearly_labels),
        'yearly_data': json.dumps(yearly_data),
        'usr': request.user,
        'cmp': cmp
    })


#===========================================START CREDITNOTE===========================================       

def get_partydetails11(request):
    if request.method == 'POST':
        if request.user.is_company:
            cmp = request.user.company
        else:
            cmp = request.user.employee.company

        try:
            party_id = request.POST.get('id').split(" ")[0]
            party = get_object_or_404(Party, company=cmp, id=party_id)

            balance = party.openingbalance if party.openingbalance else None
            phone = party.contact if party.contact else None
            address = party.address if party.address else None
            payment1 = party.payment if party.payment else None

            # Initialize lists to store multiple bill numbers and dates
            invoiceno = []
            invoicedate = []

            try:
                # Retrieve all instances for the party
                invoice_instances = Invoice.objects.filter(party=party, company=cmp)

                # Get invoice numbers already saved in the CreditNote table
                creditnote_invoice_numbers = CreditNote.objects.filter(salesinvoice__in=invoice_instances).values_list('salesinvoice__invoice_no', flat=True)

                # Loop through each instance and collect bill numbers and dates, excluding those in CreditNote
                for bill_instance in invoice_instances:
                    if bill_instance.invoice_no not in creditnote_invoice_numbers:
                        invoiceno.append(bill_instance.invoice_no)
                        invoicedate.append(bill_instance.invoice_date)

            except Invoice.DoesNotExist:
                pass

            # Return a JSON response with the list of bill numbers and dates
            if not invoiceno and not invoicedate:
                return JsonResponse({
                    'invoiceno': ['no invoice'],
                    'invoicedate': ['nodate'],
                    'address': ['no address'],
                    'payment1': ['no payment'],
                    'phone': ['nophone']
                })

            return JsonResponse({
                'invoiceno': invoiceno,
                'invoicedate': invoicedate,
                'address': address,
                'bal': balance,
                'phone': phone,
                'payment1': payment1,
                'id': party.id
            })

        except KeyError:
            return JsonResponse({'error': 'The key "id" is missing in the POST request.'})

        except Party.DoesNotExist:
            return JsonResponse({'error': 'Party not found.'})


def get_invoice_item1(request):
    invoice_no = request.GET.get('invoiceno')
    
    if invoice_no:
        try:
            invoice = Invoice.objects.get(invoice_no=invoice_no)
            invoice_items = InvoiceItem.objects.filter(invoice=invoice)

            items = []
            for item in invoice_items:
                items.append({
                    'id': item.items.id,
                    'name': item.item,
                    'hsn': item.hsn,
                    'quantity': item.quantity,
                    'price': item.price,
                    'tax': item.tax,
                    'discount': item.discount,
                    'total': item.total
                })

            return JsonResponse({'items': items}, status=200)
        except Invoice.DoesNotExist:
            return JsonResponse({'error': 'Invoice not found'}, status=404)
    else:
        return JsonResponse({'error': 'No invoice number provided'}, status=400)

 #==========================================START DEBITNOTE========================================
 #--------------------------------------UPDATE--------------------------------------   

def get_bill_item(request):
    user_id = request.user.id  # Assuming user is authenticated and user_id is obtained from the request
    print('5:',user_id)
    billno = request.GET.get('billno', None)
    print('6:',billno)
    
    if not billno:
        return JsonResponse({'error': 'Bill number not provided'}, status=400)
    
    try:
        # Query the PurchaseBill with the specified billno and user_id
        bill = PurchaseBill.objects.get(billno=billno, staff_id=user_id)
        print('7:',bill)
        items = PurchaseBillItem.objects.filter(purchasebill=bill)
        print('8:',items)
        
        items_data = [
            {
                'id': item.product.id,
                'name': item.item,
                'hsn': item.hsn,
                'quantity': item.qty,
                'price': item.price,
                'tax': item.VAT,
                'discount': item.discount,
                'total': item.total,
            }
            for item in items
        ]
        return JsonResponse({'items': items_data})
    except PurchaseBill.DoesNotExist:
        return JsonResponse({'error': 'Bill not found or you do not have access to this bill'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)


def history_debit(request,id):
  if request.user.is_company:
    cmp = request.user.company
  else:
    cmp = request.user.employee.company 
  usr = CustomUser.objects.get(username=request.user) 
  dbill = DebitNote.objects.get(id=id)
  hst= DebitNoteHistory.objects.filter(debit_note=dbill,company=cmp)

  context = {'hst':hst,'dbill':dbill,'usr':request.user}
  return render(request,'debithistory.html',context)

def savePartycredit(request):
    if request.user.is_authenticated:
      if request.method == "POST":
          if request.user.is_company:
            cmp = request.user.company
          else:
            cmp = request.user.employee.company  
          print(cmp)
          usr = CustomUser.objects.get(username=request.user)
          party_name = request.POST['party_name']
          gst_no = request.POST['gst_no']
          mob = request.POST['party_num']
          gsttype = request.POST['gsttype']
          #state = request.POST['state']
          email = request.POST['email']
          addr = request.POST['party_addr']
          opbal = request.POST['creditamt']
          payment = request.POST.get('paymentType','')
          date = request.POST['credit_date']
          add1 = request.POST['addField1']
          add2 = request.POST['addField2']
          add3 = request.POST['addField3']
          if Party.objects.filter(trn_no=gst_no, company=cmp).exists() and not(gsttype == 'Unregistered/Consumers'):
            error_message = 'TRN number already exists!!!'
            return JsonResponse({'error_message': error_message})
          if gst_no == '' and (gsttype == 'Registered Business - Regular' or gsttype == 'Registered Business - Composition'):
            error_message = 'TRN number required.'
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(email=email, company=cmp).exists():
            error_message = 'Email already exists!!!'
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(contact=mob,company=cmp).exists():
            error_message = 'Mobile number already exists!!!'
            return JsonResponse({'error_message': error_message})
          user = request.user  
          party = Party(
                  user=user,
                  company=cmp,
                  party_name=party_name,
                  trn_no=gst_no,
                  contact=mob,
                  trn_type=gsttype,
                  #state=state,
                  address=addr,
                  email=email,
                  openingbalance=opbal,
                  payment=payment,
                  current_date=date,
                  additionalfield1=add1,
                  additionalfield2=add2,
                  additionalfield3=add3
              )
          party.save()
          print('Party created succefully ')
          trans = Transactions_party(user=request.user,company=cmp, trans_type='Opening Balance', trans_number=gst_no,
                                           trans_date=date, total=opbal, balance=opbal, party=party)
          trans.save()
          tr_history = PartyTransactionHistory(party=party, Transactions_party=trans, action="CREATED")
          tr_history.save()
          return redirect('SalesReturn')
          
          
def savePartydebit(request):
    if request.user.is_authenticated:
      if request.method == "POST":
          if request.user.is_company:
            cmp = request.user.company
          else:
            cmp = request.user.employee.company  
          print(cmp)
          usr = CustomUser.objects.get(username=request.user)
          party_name = request.POST['party_name']
          gst_no = request.POST['gst_no']
          mob = request.POST['party_num']
          gsttype = request.POST['gsttype']
          #state = request.POST['state']
          email = request.POST['email']
          addr = request.POST['party_addr']
          opbal = request.POST['creditamt']
          payment = request.POST.get('paymentType','')
          date = request.POST['credit_date']
          add1 = request.POST['addField1']
          add2 = request.POST['addField2']
          add3 = request.POST['addField3']
          if Party.objects.filter(trn_no=gst_no, company=cmp).exists() and not(gsttype == 'Unregistered/Consumers'):
            error_message = 'TRN number already exists!!!'
            return JsonResponse({'error_message': error_message})
          if gst_no == '' and (gsttype == 'Registered Business - Regular' or gsttype == 'Registered Business - Composition'):
            error_message = 'TRN number required.'
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(email=email, company=cmp).exists():
            error_message = 'Email already exists!!!'
            return JsonResponse({'error_message': error_message})
          if Party.objects.filter(contact=mob,company=cmp).exists():
            error_message = 'Mobile number already exists!!!'
            return JsonResponse({'error_message': error_message})
          user = request.user  
          party = Party(
                  user=user,
                  company=cmp,
                  party_name=party_name,
                  trn_no=gst_no,
                  contact=mob,
                  trn_type=gsttype,
                  #state=state,
                  address=addr,
                  email=email,
                  openingbalance=opbal,
                  payment=payment,
                  current_date=date,
                  additionalfield1=add1,
                  additionalfield2=add2,
                  additionalfield3=add3
              )
          party.save()
          print('Party created succefully ')
          trans = Transactions_party(user=request.user,company=cmp, trans_type='Opening Balance', trans_number=gst_no,
                                           trans_date=date, total=opbal, balance=opbal, party=party)
          trans.save()
          tr_history = PartyTransactionHistory(party=party, Transactions_party=trans, action="CREATED")
          tr_history.save()
          return redirect('createdebitnote')
          
          
def save_partydebit(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company  
    usr = CustomUser.objects.get(username=request.user)

    if request.method == 'POST':
        partyname = request.POST.get('partyname')
        trn_no = request.POST.get('trn_no')
        contact = request.POST.get('contact')
        placeof=request.POST.get('pol')
        trn_type = request.POST.get('trn_type')
        address = request.POST.get('address')
        email = request.POST.get('email')
        balance = request.POST.get('balance')
        stock=request.POST.get('stock')
        paymentType = request.POST.get('paymentType')
        currentdate = request.POST.get('currentdate')
        additionalfield1 = request.POST.get('additionalfield1')
        additionalfield2 = request.POST.get('additionalfield2')
        additionalfield3 = request.POST.get('additionalfield3')
        # print(trn_no)
        # print(partyname)

        # Check if the contact number already exists in the database
        if Party.objects.filter(contact=contact,company=cmp).exists():
            # Send a message indicating that the contact number already exists

            return redirect('createdebitnote')

        # Check if the transaction number already exists in the database
        if trn_type == "Unregistered/Consumers":
            Party.objects.create(    user=usr,
            company=cmp,
            party_name=partyname,
            trn_no=trn_no,
            contact=contact,
            trn_type=trn_type,
            address=address,
            email=email,
            state=placeof,
            openingbalance=balance,
            opening_stock=stock,
            payment=paymentType,
            current_date=currentdate,
            additionalfield1=additionalfield1,
            additionalfield2=additionalfield2,
            additionalfield3=additionalfield3)
            # Optionally, you can send a success message here

            return redirect('createdebitnote')
        else:
             if Party.objects.filter(trn_no=trn_no, company=cmp).exists():
                # Send a message indicating that the transaction number already exists

                return redirect('createdebitnote')
        Party.objects.create(
            user=usr,
            company=cmp,
            party_name=partyname,
            trn_no=trn_no,
            contact=contact,
            trn_type=trn_type,
            address=address,
            email=email,
            state=placeof,
            openingbalance=balance,
            opening_stock=stock,
            payment=paymentType,
            current_date=currentdate,
            additionalfield1=additionalfield1,
            additionalfield2=additionalfield2,
            additionalfield3=additionalfield3
        )
        return redirect('createdebitnote')